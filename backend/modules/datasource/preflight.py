import uuid
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class ExcelPreflight:
    temp_path: Path
    sheets: list[str]
    tables: dict[str, list[str]]
    named_ranges: list[str]
    created_at: datetime
    delete_file: bool


_PREFLIGHTS: dict[str, ExcelPreflight] = {}
_PREFLIGHT_TTL = timedelta(minutes=30)


def create_preflight(file_path: Path, *, delete_file: bool = True) -> tuple[str, ExcelPreflight]:
    workbook = load_workbook(file_path, read_only=False, data_only=True)
    sheets = workbook.sheetnames
    tables: dict[str, list[str]] = {}
    for sheet in workbook.worksheets:
        if not hasattr(sheet, 'tables') or not sheet.tables:
            continue
        tables[sheet.title] = list(sheet.tables.keys())

    named_ranges = [name for name in workbook.defined_names]

    preflight_id = str(uuid.uuid4())
    preflight = ExcelPreflight(
        temp_path=file_path,
        sheets=sheets,
        tables=tables,
        named_ranges=named_ranges,
        created_at=datetime.now(UTC).replace(tzinfo=None),
        delete_file=delete_file,
    )
    _PREFLIGHTS[preflight_id] = preflight
    return preflight_id, preflight


def get_preflight(preflight_id: str) -> ExcelPreflight | None:
    _cleanup_expired()
    return _PREFLIGHTS.get(preflight_id)


def clear_preflight(preflight_id: str, *, delete_file: bool = True) -> None:
    preflight = _PREFLIGHTS.pop(preflight_id, None)
    if not preflight:
        return
    if delete_file and preflight.temp_path.exists():
        preflight.temp_path.unlink()


def _cleanup_expired() -> None:
    now = datetime.now(UTC).replace(tzinfo=None)
    expired = [
        (preflight_id, preflight.delete_file)
        for preflight_id, preflight in _PREFLIGHTS.items()
        if now - preflight.created_at > _PREFLIGHT_TTL
    ]
    for preflight_id, delete_file in expired:
        clear_preflight(preflight_id, delete_file=delete_file)
