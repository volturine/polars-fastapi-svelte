import logging
import os
import shutil
import sqlite3
import uuid
from collections.abc import Callable
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import polars as pl
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from pyiceberg.catalog import load_catalog
from pyiceberg.table import Table
from sqlalchemy import select, update
from sqlmodel import Session, col

from core.exceptions import DataSourceConnectionError, DataSourceNotFoundError, DataSourceValidationError, FileError
from core.namespace import get_namespace, namespace_paths
from modules.compute.operations.datasource import load_datasource
from modules.datasource.models import DataSource
from modules.datasource.schemas import (
    ColumnSchema,
    ColumnStats,
    ColumnStatsResponse,
    CSVOptions,
    DataSourceResponse,
    DataSourceUpdate,
    FileListItem,
    FileListResponse,
    SchemaDiff,
    SchemaInfo,
    SnapshotCompareResponse,
    SnapshotPreview,
)
from modules.datasource.source_types import FILE_BASED_CATEGORIES, SOURCE_TYPE_CATEGORY, DataSourceType
from modules.engine_runs import service as engine_run_service

logger = logging.getLogger(__name__)


def _prepare_clean_target(clean_dir: Path, datasource_id: str, branch: str) -> Path:
    target = clean_dir / datasource_id / branch
    target.mkdir(parents=True, exist_ok=True)
    return target


def _write_iceberg_table(lazy: pl.LazyFrame, table_path: Path, build_mode: str) -> Table:
    catalog_path = table_path.parent / 'catalog.db'
    if not catalog_path.exists():
        catalog_path.touch()
    catalog_config = {
        'type': 'sql',
        'uri': f'sqlite:///{catalog_path}',
        'warehouse': f'file://{table_path.parent}',
    }
    catalog = load_catalog('local', **catalog_config)
    namespace = 'clean'
    catalog.create_namespace_if_not_exists(namespace)
    identifier = f'{namespace}.{table_path.parent.name}'
    arrow_table = lazy.collect().to_arrow()
    if build_mode == 'recreate' and catalog.table_exists(identifier):
        catalog.drop_table(identifier)
    if catalog.table_exists(identifier):
        table = catalog.load_table(identifier)
        if build_mode == 'incremental':
            table.append(arrow_table)
        else:
            table.overwrite(arrow_table)
        return table
    table = catalog.create_table(identifier, schema=arrow_table.schema, location=str(table_path))
    table.append(arrow_table)
    return table


def _build_iceberg_config(paths, target_path: Path, branch: str, source_config: dict | None = None) -> dict:
    catalog_path = target_path.parent / 'catalog.db'
    return {
        'catalog_type': 'sql',
        'catalog_uri': f'sqlite:///{catalog_path}',
        'warehouse': f'file://{paths.clean_dir}',
        'namespace': 'clean',
        'table': target_path.parent.name,
        'metadata_path': str(target_path.parent),
        'branch': branch,
        'source': source_config,
        'namespace_name': get_namespace(),
        'refresh': None,
    }


def create_file_datasource(
    session: Session,
    name: str,
    file_path: str,
    file_type: str,
    options: dict | None = None,
    csv_options: CSVOptions | None = None,
    sheet_name: str | None = None,
    start_row: int | None = None,
    start_col: int | None = None,
    end_col: int | None = None,
    end_row: int | None = None,
    has_header: bool | None = None,
    table_name: str | None = None,
    named_range: str | None = None,
    cell_range: str | None = None,
) -> DataSourceResponse:
    datasource_id = str(uuid.uuid4())
    resolved_path = Path(os.path.realpath(Path(file_path).resolve()))
    paths = namespace_paths()
    data_root = Path(os.path.realpath(paths.base_dir.resolve()))
    upload_root = Path(os.path.realpath(paths.upload_dir.resolve()))
    within_data_root = data_root in resolved_path.parents or data_root == resolved_path
    within_upload_root = upload_root in resolved_path.parents or upload_root == resolved_path
    if not (within_data_root or within_upload_root):
        raise ValueError(f'Path must be inside data directory: {data_root}')
    if file_type in {'csv', 'json', 'ndjson', 'excel'} and not resolved_path.is_file():
        raise ValueError(f'Path must be a file for type: {file_type}')
    if file_type == 'parquet' and not (resolved_path.is_file() or resolved_path.is_dir()):
        raise ValueError('Parquet path must be a file or directory')

    file_config = {
        'file_path': str(resolved_path),
        'file_type': file_type,
        'options': options or {},
        'csv_options': csv_options.model_dump() if csv_options else None,
        'sheet_name': sheet_name,
        'start_row': start_row,
        'start_col': start_col,
        'end_col': end_col,
        'end_row': end_row,
        'has_header': has_header,
        'table_name': table_name,
        'named_range': named_range,
        'cell_range': cell_range,
    }
    config = {'source_type': DataSourceType.FILE, **file_config}
    try:
        lazy = load_datasource(config)
    except Exception as exc:
        raise DataSourceValidationError(
            f'Failed to load file datasource for ingestion: {exc}',
            details={'file_path': str(resolved_path), 'file_type': file_type},
        ) from exc
    target_path = _prepare_clean_target(paths.clean_dir, datasource_id, 'master')
    snapshot = _write_iceberg_table(lazy, target_path, build_mode='recreate')
    iceberg_config = _build_iceberg_config(paths, target_path, branch='master', source_config=config)
    current_snapshot = snapshot.current_snapshot() if snapshot else None
    if current_snapshot:
        snapshot_id = current_snapshot.snapshot_id
        snapshot_ts = current_snapshot.timestamp_ms
        iceberg_config['snapshot_id'] = str(snapshot_id)
        iceberg_config['snapshot_timestamp_ms'] = int(snapshot_ts)
    datasource = DataSource(
        id=datasource_id,
        name=name,
        source_type=DataSourceType.ICEBERG,
        config=iceberg_config,
        created_at=datetime.now(UTC).replace(tzinfo=None),
    )

    session.add(datasource)
    session.commit()
    session.refresh(datasource)

    _log_datasource_create(session, datasource_id, name, DataSourceType.ICEBERG, iceberg_config, branch='master')

    try:
        schema_info = _extract_schema(datasource)
        datasource.schema_cache = schema_info.model_dump()
        session.commit()
        session.refresh(datasource)
        logger.info(f'Cached schema for datasource {datasource_id}: {len(schema_info.columns)} columns, {schema_info.row_count} rows')
    except Exception as e:
        session.rollback()
        raise DataSourceValidationError(
            f'Failed to extract schema for datasource {datasource_id}: {e}',
            details={'datasource_id': datasource_id},
        ) from e

    logger.info(f'Created iceberg datasource {datasource_id} ({name}) from file {file_path}')
    response = DataSourceResponse.model_validate(datasource)
    response.config['source'] = file_config
    response.config.update(file_config)
    return response


def create_placeholder_output_datasource(
    session: Session,
    result_id: str,
    analysis_id: str,
    analysis_tab_id: str,
) -> None:
    try:
        uuid.UUID(result_id)
    except ValueError:
        raise ValueError(f'result_id must be a valid UUID, got: {result_id!r}') from None
    existing = session.get(DataSource, result_id)
    if existing:
        return
    datasource = DataSource(
        id=result_id,
        name=result_id,
        source_type=DataSourceType.ANALYSIS,
        config={'analysis_tab_id': analysis_tab_id},
        created_by_analysis_id=analysis_id,
        created_by='analysis',
        is_hidden=False,
        created_at=datetime.now(UTC).replace(tzinfo=None),
    )
    session.add(datasource)
    session.flush()


def create_analysis_datasource(
    session: Session,
    name: str,
    analysis_id: str,
    analysis_tab_id: str | None = None,
    is_hidden: bool = False,
    source_type: DataSourceType = DataSourceType.ANALYSIS,
) -> DataSourceResponse:
    from modules.analysis.models import Analysis

    analysis = session.get(Analysis, analysis_id)
    if not analysis:
        raise ValueError(f'Analysis {analysis_id} not found')
    datasource_id = str(uuid.uuid4())
    config = {}
    if analysis_tab_id:
        config['analysis_tab_id'] = analysis_tab_id

    datasource = DataSource(
        id=datasource_id,
        name=name,
        source_type=source_type,
        config=config,
        created_by_analysis_id=analysis_id,
        created_by='analysis',
        is_hidden=is_hidden,
        created_at=datetime.now(UTC).replace(tzinfo=None),
    )

    session.add(datasource)
    session.commit()
    session.refresh(datasource)

    _log_datasource_create(session, datasource_id, name, source_type, config, branch='master')
    return DataSourceResponse.model_validate(datasource)


@dataclass
class ExcelPreviewResult:
    preview: list[list[str | None]]
    detected_end_row: int | None
    sheet_name: str
    start_row: int
    start_col: int
    end_col: int
    end_row: int


def build_excel_preview(
    file_path: Path,
    sheet_name: str,
    start_row: int,
    start_col: int,
    end_col: int,
    end_row: int | None,
    has_header: bool,
    table_name: str | None = None,
    named_range: str | None = None,
    cell_range: str | None = None,
    preview_rows: int = 100,
) -> ExcelPreviewResult:
    workbook = load_workbook(file_path, read_only=False, data_only=True)
    resolved = _resolve_excel_bounds(
        workbook,
        sheet_name,
        start_row,
        start_col,
        end_col,
        end_row,
        table_name,
        named_range,
        cell_range,
    )
    sheet = workbook[resolved.sheet_name]

    end_row_value = resolved.end_row
    if end_row_value is None:
        end_row_value = _detect_end_row(sheet, resolved.start_row, resolved.start_col, resolved.end_col)
    _validate_excel_bounds(sheet, resolved.start_row, resolved.start_col, resolved.end_col, end_row_value)

    preview_end_row = min(resolved.start_row + preview_rows - 1, end_row_value)
    rows = _collect_preview_rows(sheet, resolved.start_row, resolved.start_col, resolved.end_col, preview_end_row)
    return ExcelPreviewResult(
        preview=rows,
        detected_end_row=end_row_value,
        sheet_name=resolved.sheet_name,
        start_row=resolved.start_row,
        start_col=resolved.start_col,
        end_col=resolved.end_col,
        end_row=end_row_value,
    )


def resolve_excel_selection(
    file_path: Path,
    sheet_name: str | None,
    start_row: int,
    start_col: int,
    end_col: int,
    end_row: int | None,
    table_name: str | None = None,
    named_range: str | None = None,
    cell_range: str | None = None,
) -> tuple[str, int, int, int, int]:
    workbook = load_workbook(file_path, read_only=False, data_only=True)
    target_sheet = sheet_name or (workbook.sheetnames[0] if workbook.sheetnames else None)
    if not target_sheet:
        raise ValueError('No sheets found in file')
    resolved = _resolve_excel_bounds(
        workbook,
        target_sheet,
        start_row,
        start_col,
        end_col,
        end_row,
        table_name,
        named_range,
        cell_range,
    )
    sheet = workbook[resolved.sheet_name]
    end_row_value = resolved.end_row
    if end_row_value is None:
        end_row_value = _detect_end_row(sheet, resolved.start_row, resolved.start_col, resolved.end_col)
    _validate_excel_bounds(sheet, resolved.start_row, resolved.start_col, resolved.end_col, end_row_value)
    return resolved.sheet_name, resolved.start_row, resolved.start_col, resolved.end_col, end_row_value


@dataclass
class _ExcelBounds:
    sheet_name: str
    start_row: int
    start_col: int
    end_col: int
    end_row: int | None


def _resolve_excel_bounds(
    workbook,
    sheet_name: str,
    start_row: int,
    start_col: int,
    end_col: int,
    end_row: int | None,
    table_name: str | None,
    named_range: str | None,
    cell_range: str | None,
) -> _ExcelBounds:
    if table_name:
        sheet = workbook[sheet_name]
        tables = getattr(sheet, 'tables', None)
        if not tables:
            raise ValueError(f'No tables available in sheet: {sheet_name}')
        table = tables.get(table_name)
        if not table:
            raise ValueError(f'Table not found: {table_name}')
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if min_col is None or min_row is None or max_col is None or max_row is None:
            raise ValueError(f'Invalid table range: {table_name}')
        min_col = int(min_col)
        min_row = int(min_row)
        max_col = int(max_col)
        max_row = int(max_row)
        return _ExcelBounds(sheet_name, min_row - 1, min_col - 1, max_col - 1, max_row - 1)

    if named_range:
        defined = workbook.defined_names.get(named_range)
        if not defined:
            raise ValueError(f'Named range not found: {named_range}')
        destinations = list(defined.destinations)
        if not destinations:
            raise ValueError(f'Named range has no destinations: {named_range}')
        dest_sheet, coord = destinations[0]
        min_col, min_row, max_col, max_row = range_boundaries(coord)
        if min_col is None or min_row is None or max_col is None or max_row is None:
            raise ValueError(f'Invalid named range: {named_range}')
        min_col = int(min_col)
        min_row = int(min_row)
        max_col = int(max_col)
        max_row = int(max_row)
        return _ExcelBounds(dest_sheet, min_row - 1, min_col - 1, max_col - 1, max_row - 1)

    if cell_range:
        return _parse_cell_range(workbook, cell_range, sheet_name)

    resolved_start_row = max(start_row, 0)
    resolved_start_col = max(start_col, 0)
    resolved_end_col = end_col
    if resolved_end_col <= resolved_start_col:
        sheet = workbook[sheet_name]
        resolved_end_col = _detect_end_col(sheet, resolved_start_row, resolved_start_col)
    resolved_end_col = max(resolved_end_col, resolved_start_col)
    end_row_value = end_row
    if end_row_value is not None:
        end_row_value = max(end_row_value, resolved_start_row)
    return _ExcelBounds(sheet_name, resolved_start_row, resolved_start_col, resolved_end_col, end_row_value)


def _parse_cell_range(workbook, cell_range: str, default_sheet: str | None) -> _ExcelBounds:
    raw = cell_range.strip()
    if not raw:
        raise ValueError('Cell range cannot be empty')
    target_sheet = default_sheet
    coord = raw
    if '!' in raw:
        sheet_part, coord_part = raw.split('!', maxsplit=1)
        sheet_part = sheet_part.strip()
        if sheet_part.startswith("'") and sheet_part.endswith("'"):
            sheet_part = sheet_part[1:-1]
        if not sheet_part:
            raise ValueError(f'Invalid cell range sheet: {cell_range}')
        target_sheet = sheet_part
        coord = coord_part.strip()
    if not target_sheet:
        target_sheet = workbook.sheetnames[0] if workbook.sheetnames else None
    if not target_sheet or target_sheet not in workbook.sheetnames:
        raise ValueError(f'Sheet not found for cell range: {target_sheet}')
    min_col, min_row, max_col, max_row = range_boundaries(coord)
    if min_col is None or min_row is None or max_col is None or max_row is None:
        raise ValueError(f'Invalid cell range: {cell_range}')
    return _ExcelBounds(target_sheet, int(min_row) - 1, int(min_col) - 1, int(max_col) - 1, int(max_row) - 1)


def format_excel_cell_range(
    sheet_name: str,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
) -> str:
    start_cell = f'{get_column_letter(start_col + 1)}{start_row + 1}'
    end_cell = f'{get_column_letter(end_col + 1)}{end_row + 1}'
    return f'{sheet_name}!{start_cell}:{end_cell}'


def _validate_excel_bounds(sheet, start_row: int, start_col: int, end_col: int, end_row: int) -> None:
    if start_row < 0 or start_col < 0:
        raise ValueError('Excel bounds must be non-negative')
    if end_row < start_row or end_col < start_col:
        raise ValueError('Excel bounds are invalid')
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0
    if max_row <= 0 or max_col <= 0:
        raise ValueError('Excel sheet has no data')
    if start_row >= max_row or end_row >= max_row:
        raise ValueError('Excel row bounds exceed sheet size')
    if start_col >= max_col or end_col >= max_col:
        raise ValueError('Excel column bounds exceed sheet size')


def _detect_end_col(sheet, start_row: int, start_col: int) -> int:
    """Scan the header row (start_row) rightward to find the last non-empty column."""
    max_col = sheet.max_column or 0
    if max_col <= start_col:
        return start_col
    last_col = start_col
    for cell in sheet.iter_rows(
        min_row=start_row + 1,
        max_row=start_row + 1,
        min_col=start_col + 1,
        max_col=max_col,
    ):
        for c in cell:
            if c.value is not None and str(c.value).strip() != '':
                last_col = c.column - 1  # 0-indexed
    return last_col


def _detect_end_row(sheet, start_row: int, start_col: int, end_col: int) -> int:
    max_row = sheet.max_row or 0
    if max_row <= start_row:
        return start_row
    for row_index in range(start_row + 1, max_row + 1):
        values: list[object | None] = []
        for cell in sheet.iter_rows(min_row=row_index, max_row=row_index, min_col=start_col + 1, max_col=end_col + 1):
            values = [c.value for c in cell]
        if all(value is None or str(value).strip() == '' for value in values):
            return max(start_row, row_index - 2)
    return max_row - 1


def _collect_preview_rows(
    sheet,
    start_row: int,
    start_col: int,
    end_col: int,
    preview_end_row: int,
) -> list[list[str | None]]:
    rows: list[list[str | None]] = []
    for row in sheet.iter_rows(
        min_row=start_row + 1,
        max_row=preview_end_row + 1,
        min_col=start_col + 1,
        max_col=end_col + 1,
    ):
        values = [cell.value for cell in row]
        rows.append([str(value) if value is not None else None for value in values])
    return rows


def create_database_datasource(
    session: Session,
    name: str,
    connection_string: str,
    query: str,
    branch: str = 'master',
) -> DataSourceResponse:
    datasource_id = str(uuid.uuid4())
    branch_name = branch
    source_config = {
        'connection_string': connection_string,
        'query': query,
        'branch': branch_name,
    }
    try:
        lazy = load_datasource(
            {
                'source_type': DataSourceType.DATABASE,
                'connection_string': connection_string,
                'query': query,
            }
        )
    except Exception as exc:
        if connection_string.startswith('postgresql://'):
            datasource = DataSource(
                id=datasource_id,
                name=name,
                source_type=DataSourceType.DATABASE,
                config=source_config,
                created_at=datetime.now(UTC).replace(tzinfo=None),
            )
            session.add(datasource)
            session.commit()
            session.refresh(datasource)
            _log_datasource_create(session, datasource_id, name, DataSourceType.DATABASE, source_config, branch=branch_name)
            return DataSourceResponse.model_validate(datasource)
        raise DataSourceConnectionError(
            'Failed to query database datasource',
            details={'connection_string': connection_string},
        ) from exc
    paths = namespace_paths()
    target_path = _prepare_clean_target(paths.clean_dir, datasource_id, branch_name)
    snapshot = _write_iceberg_table(lazy, target_path, build_mode='recreate')
    iceberg_config = _build_iceberg_config(paths, target_path, branch=branch_name, source_config=source_config)
    current_snapshot = snapshot.current_snapshot() if snapshot else None
    if current_snapshot:
        snapshot_id = current_snapshot.snapshot_id
        snapshot_ts = current_snapshot.timestamp_ms
        iceberg_config['snapshot_id'] = str(snapshot_id)
        iceberg_config['snapshot_timestamp_ms'] = int(snapshot_ts)

    datasource = DataSource(
        id=datasource_id,
        name=name,
        source_type=DataSourceType.ICEBERG,
        config=iceberg_config,
        created_at=datetime.now(UTC).replace(tzinfo=None),
    )

    session.add(datasource)
    session.commit()
    session.refresh(datasource)

    _log_datasource_create(session, datasource_id, name, DataSourceType.ICEBERG, iceberg_config, branch=branch_name)
    return DataSourceResponse.model_validate(datasource)


def create_iceberg_datasource(
    session: Session,
    name: str,
    source: dict,
    branch: str = 'master',
) -> DataSourceResponse:
    source_type = source.get('source_type') if isinstance(source, dict) else None
    if source_type not in {DataSourceType.FILE, DataSourceType.DATABASE}:
        raise DataSourceValidationError(
            'Iceberg datasource source_type is not supported for ingestion',
            details={'source_type': source_type},
        )
    if source_type == DataSourceType.DATABASE:
        connection_string = source.get('connection_string')
        query = source.get('query')
        if not connection_string or not query:
            raise DataSourceValidationError(
                'Datasource source is missing connection details',
                details={'source_type': source_type},
            )
        try:
            lazy = load_datasource(
                {
                    'source_type': DataSourceType.DATABASE,
                    'connection_string': connection_string,
                    'query': query,
                }
            )
        except Exception as exc:
            raise DataSourceConnectionError(
                'Failed to query database datasource',
                details={'source_type': source_type},
            ) from exc
    else:
        try:
            lazy = load_datasource(source)
        except Exception as exc:
            raise DataSourceConnectionError(
                'Failed to read file datasource',
                details={'source_type': source_type},
            ) from exc

    datasource_id = str(uuid.uuid4())
    paths = namespace_paths()
    branch_name = branch or 'master'
    target_path = _prepare_clean_target(paths.clean_dir, datasource_id, branch_name)
    snapshot = _write_iceberg_table(lazy, target_path, build_mode='recreate')
    config = _build_iceberg_config(paths, target_path, branch=branch_name, source_config=source)
    current_snapshot = snapshot.current_snapshot() if snapshot else None
    if current_snapshot:
        config['snapshot_id'] = str(current_snapshot.snapshot_id)
        config['snapshot_timestamp_ms'] = int(current_snapshot.timestamp_ms)

    datasource = DataSource(
        id=datasource_id,
        name=name,
        source_type=DataSourceType.ICEBERG,
        config=config,
        created_at=datetime.now(UTC).replace(tzinfo=None),
    )

    session.add(datasource)
    session.commit()
    session.refresh(datasource)

    _log_datasource_create(session, datasource_id, name, DataSourceType.ICEBERG, config, branch=branch_name)

    try:
        schema_info = _extract_schema(datasource)
        datasource.schema_cache = schema_info.model_dump()
        session.commit()
        session.refresh(datasource)
    except Exception as e:
        session.rollback()
        raise DataSourceValidationError(
            f'Failed to extract schema for datasource {datasource_id}: {e}',
            details={'datasource_id': datasource_id},
        ) from e

    logger.info(f'Created Iceberg datasource {datasource_id} ({name}) from source {source_type}')
    return DataSourceResponse.model_validate(datasource)


def refresh_external_datasource(session: Session, datasource_id: str) -> DataSourceResponse:
    datasource = session.get(DataSource, datasource_id)
    if not datasource:
        raise DataSourceNotFoundError(datasource_id)
    if datasource.source_type != DataSourceType.ICEBERG:
        raise DataSourceValidationError(
            'Refresh is only available for Iceberg datasources',
            details={'datasource_id': datasource_id},
        )
    source = datasource.config.get('source') if isinstance(datasource.config, dict) else None
    if not isinstance(source, dict):
        raise DataSourceValidationError(
            'Datasource has no external source configuration',
            details={'datasource_id': datasource_id},
        )
    source_type = source.get('source_type')
    if source_type not in {DataSourceType.DATABASE, DataSourceType.FILE}:
        raise DataSourceValidationError(
            'Datasource source is not refreshable',
            details={'datasource_id': datasource_id, 'source_type': source_type},
        )
    branch_raw = datasource.config.get('branch', source.get('branch', 'master'))
    branch = branch_raw if isinstance(branch_raw, str) and branch_raw else 'master'
    if source_type == DataSourceType.DATABASE:
        connection_string = source.get('connection_string')
        query = source.get('query')
        if not connection_string or not query:
            raise DataSourceValidationError(
                'Datasource source is missing connection details',
                details={'datasource_id': datasource_id},
            )
        try:
            lazy = load_datasource(
                {
                    'source_type': DataSourceType.DATABASE,
                    'connection_string': connection_string,
                    'query': query,
                }
            )
        except Exception as exc:
            raise DataSourceConnectionError(
                'Failed to query database datasource',
                details={'datasource_id': datasource_id},
            ) from exc
    else:
        try:
            lazy = load_datasource(source)
        except Exception as exc:
            raise DataSourceConnectionError(
                'Failed to read file datasource',
                details={'datasource_id': datasource_id},
            ) from exc
    metadata_path = datasource.config.get('metadata_path')
    if not metadata_path:
        raise DataSourceValidationError(
            'Datasource missing metadata_path',
            details={'datasource_id': datasource_id},
        )
    paths = namespace_paths()
    target_path = Path(metadata_path)
    if target_path.name != branch:
        target_path = _prepare_clean_target(paths.clean_dir, datasource_id, branch)
    snapshot = _write_iceberg_table(lazy, target_path, build_mode='recreate')

    current_snapshot = snapshot.current_snapshot() if snapshot else None
    next_config = dict(datasource.config)
    if current_snapshot:
        snapshot_id = current_snapshot.snapshot_id
        snapshot_ts = current_snapshot.timestamp_ms
        next_config['snapshot_id'] = str(snapshot_id)
        next_config['snapshot_timestamp_ms'] = int(snapshot_ts)
    next_config['branch'] = branch
    next_config['metadata_path'] = str(target_path)
    next_config['source'] = source
    next_config['refresh'] = {
        'refreshed_at': datetime.now(UTC).replace(tzinfo=None).isoformat(),
    }
    datasource.config = next_config
    datasource.schema_cache = None
    session.add(datasource)
    session.commit()
    session.refresh(datasource)

    _log_datasource_update(session, datasource.id, datasource.name, datasource.config, branch=branch)
    return DataSourceResponse.model_validate(datasource)


def refresh_datasource_for_schedule(session: Session, datasource_id: str) -> DataSourceResponse:
    datasource = session.get(DataSource, datasource_id)
    if not datasource:
        raise DataSourceNotFoundError(datasource_id)

    if is_reingestable_raw_datasource(datasource):
        return refresh_external_datasource(session, datasource_id)

    schema = _extract_schema(datasource)
    next_config = dict(datasource.config) if isinstance(datasource.config, dict) else {}
    next_config['refresh'] = {
        'refreshed_at': datetime.now(UTC).replace(tzinfo=None).isoformat(),
        'mode': 'schedule_schema_refresh',
    }

    datasource.config = next_config
    datasource.schema_cache = schema.model_dump()
    session.add(datasource)
    session.commit()
    session.refresh(datasource)

    payload = engine_run_service.create_engine_run_payload(
        analysis_id=None,
        datasource_id=datasource.id,
        kind='datasource_update',
        status='success',
        request_json={
            'name': datasource.name,
            'source_type': datasource.source_type,
            'mode': 'schedule_schema_refresh',
            'config': datasource.config,
        },
        result_json={
            'datasource_id': datasource.id,
            'datasource_name': datasource.name,
            'schema_columns': len(schema.columns),
            'row_count': schema.row_count,
        },
        triggered_by='schedule',
    )
    engine_run_service.create_engine_run(session, payload)
    return DataSourceResponse.model_validate(datasource)


def is_reingestable_raw_datasource(datasource: DataSource) -> bool:
    if datasource.source_type != DataSourceType.ICEBERG:
        return False
    if datasource.created_by == 'analysis':
        return False
    if not isinstance(datasource.config, dict):
        return False
    source = datasource.config.get('source')
    if not isinstance(source, dict):
        return False
    source_type = source.get('source_type')
    return source_type in {DataSourceType.FILE, DataSourceType.DATABASE}


def _log_datasource_update(
    session: Session,
    datasource_id: str,
    name: str,
    config: dict,
    branch: str,
) -> None:
    payload = engine_run_service.create_engine_run_payload(
        analysis_id=None,
        datasource_id=datasource_id,
        kind='datasource_update',
        status='success',
        request_json={
            'name': name,
            'source_type': DataSourceType.ICEBERG.value,
            'config': config,
            'iceberg_options': {'branch': branch},
        },
        result_json={'datasource_id': datasource_id, 'datasource_name': name},
    )
    engine_run_service.create_engine_run(session, payload)


def _log_datasource_create(
    session: Session,
    datasource_id: str,
    name: str,
    source_type: DataSourceType,
    config: dict,
    branch: str,
) -> None:
    payload = engine_run_service.create_engine_run_payload(
        analysis_id=None,
        datasource_id=datasource_id,
        kind='datasource_create',
        status='success',
        request_json={
            'name': name,
            'source_type': source_type.value,
            'config': config,
            'iceberg_options': {'branch': branch},
        },
        result_json={'datasource_id': datasource_id, 'datasource_name': name},
    )
    engine_run_service.create_engine_run(session, payload)


def get_datasource_schema(
    session: Session,
    datasource_id: str,
    sheet_name: str | None = None,
    refresh: bool = False,
) -> SchemaInfo:
    datasource = session.get(DataSource, datasource_id)

    if not datasource:
        raise DataSourceNotFoundError(datasource_id)

    # Check if we have cached schema with row_count and sample_value
    if datasource.schema_cache and sheet_name is None and not refresh:
        try:
            cached = SchemaInfo.model_validate(datasource.schema_cache)
        except Exception:
            cached = None
        if cached:
            # Re-extract if row_count or sample_value is missing
            has_samples = cached.columns and any(c.sample_value is not None for c in cached.columns)
            if cached.row_count is not None and has_samples:
                logger.debug(f'Using cached schema for datasource {datasource_id}')
                return cached

    logger.info(f'Extracting schema for datasource {datasource_id}')
    if refresh and datasource.source_type == DataSourceType.ICEBERG:
        source = datasource.config.get('source') if isinstance(datasource.config, dict) else None
        if isinstance(source, dict):
            source_type = source.get('source_type')
            if source_type in {DataSourceType.DATABASE, DataSourceType.FILE}:
                refresh_external_datasource(session, datasource_id)
                datasource = session.get(DataSource, datasource_id)
                if not datasource:
                    raise DataSourceNotFoundError(datasource_id)
    schema_info = _extract_schema(datasource, sheet_name=sheet_name)

    if sheet_name is None:
        schema_cache = schema_info.model_dump()
        session.execute(
            update(DataSource)
            .where(DataSource.id == datasource_id)  # type: ignore[arg-type]
            .values(schema_cache=schema_cache)
        )
        session.commit()
        datasource.schema_cache = schema_cache

    logger.info(f'Schema extracted and cached for datasource {datasource_id}: {len(schema_info.columns)} columns')
    return schema_info


def _get_first_non_null_samples(lazy: pl.LazyFrame, max_rows: int = 1000) -> dict[str, str | None]:
    columns = lazy.collect_schema().names()
    exprs = [pl.col(col).drop_nulls().first().alias(col) for col in columns]
    result = lazy.head(max_rows).select(exprs).collect()
    if result.height == 0:
        return {col: None for col in columns}
    return {col: (str(result[col][0]) if result[col][0] is not None else None) for col in columns}


def _extract_schema(datasource: DataSource, sheet_name: str | None = None) -> SchemaInfo:
    try:
        source_type = DataSourceType(datasource.source_type)
    except ValueError as exc:
        raise DataSourceConnectionError(
            'Unsupported datasource type for schema extraction',
            details={'datasource_id': datasource.id, 'source_type': datasource.source_type},
        ) from exc
    handler = _SCHEMA_HANDLERS.get(source_type)
    if not handler:
        raise DataSourceConnectionError(
            'Unsupported datasource type for schema extraction',
            details={'datasource_id': datasource.id, 'source_type': datasource.source_type},
        )
    return handler(datasource, sheet_name)


def _schema_from_analysis(datasource: DataSource, sheet_name: str | None) -> SchemaInfo:
    raise DataSourceValidationError(
        'Schema extraction not supported for analysis datasources',
        details={'datasource_id': datasource.id},
    )


def _schema_from_database(datasource: DataSource, sheet_name: str | None) -> SchemaInfo:
    connection_string = datasource.config['connection_string']
    query = datasource.config['query']
    try:
        if connection_string.startswith('sqlite:'):
            parsed = urlparse(connection_string)
            if not parsed.path:
                raise DataSourceConnectionError('SQLite connection string must include a database path')
            connection = sqlite3.connect(parsed.path)
            try:
                frame = pl.read_database(query, connection)
            finally:
                connection.close()
        else:
            frame = pl.read_database_uri(query, connection_string)
    except Exception as e:
        raise DataSourceConnectionError(
            'Failed to query database datasource',
            details={'datasource_id': datasource.id, 'source_type': datasource.source_type},
        ) from e
    schema = frame.schema
    row_count = frame.height

    sample_values = _get_first_non_null_samples(frame.lazy())

    columns = [
        ColumnSchema(
            name=name,
            dtype=str(dtype),
            nullable=True,
            sample_value=sample_values.get(name),
        )
        for name, dtype in schema.items()
    ]

    return SchemaInfo(columns=columns, row_count=row_count)


def _schema_from_file(datasource: DataSource, sheet_name: str | None) -> SchemaInfo:
    config = {
        'source_type': datasource.source_type,
        **datasource.config,
    }
    if sheet_name:
        config = {**config, 'sheet_name': sheet_name}
    try:
        lazy = load_datasource(config)
    except Exception as e:
        category = SOURCE_TYPE_CATEGORY.get(DataSourceType(datasource.source_type))
        label = category.value if category else 'datasource'
        raise DataSourceConnectionError(
            f'Failed to load {label} datasource',
            details={'datasource_id': datasource.id, 'source_type': datasource.source_type},
        ) from e

    schema = lazy.collect_schema()
    row_count = lazy.select(pl.len()).collect().item()
    sample_values = _get_first_non_null_samples(lazy)
    columns = [
        ColumnSchema(
            name=name,
            dtype=str(dtype),
            nullable=True,
            sample_value=sample_values.get(name),
        )
        for name, dtype in schema.items()
    ]
    return SchemaInfo(columns=columns, row_count=row_count)


_SCHEMA_HANDLERS: dict[DataSourceType, Callable[[DataSource, str | None], SchemaInfo]] = {
    DataSourceType.ANALYSIS: _schema_from_analysis,
    DataSourceType.DATABASE: _schema_from_database,
    **{
        source_type: _schema_from_file
        for source_type in SOURCE_TYPE_CATEGORY
        if SOURCE_TYPE_CATEGORY[source_type] in FILE_BASED_CATEGORIES
        and source_type not in {DataSourceType.ANALYSIS, DataSourceType.DATABASE}
    },
}


def list_data_files(path: str | None) -> FileListResponse:
    base_dir = namespace_paths().base_dir.resolve()
    target = Path(path) if path else base_dir
    resolved = target.resolve()
    if base_dir not in resolved.parents and base_dir != resolved:
        raise ValueError(f'Path must be inside data directory: {base_dir}')
    if not resolved.exists():
        raise ValueError(f'Path does not exist: {resolved}')
    if not resolved.is_dir():
        raise ValueError(f'Path must be a directory: {resolved}')

    entries = [
        FileListItem(
            name=item.name,
            path=str(item),
            is_dir=item.is_dir(),
        )
        for item in sorted(resolved.iterdir(), key=lambda entry: (not entry.is_dir(), entry.name.lower()))
    ]
    return FileListResponse(base_path=str(resolved), entries=entries)


def compare_iceberg_snapshots(
    session: Session,
    datasource_id: str,
    snapshot_a: str,
    snapshot_b: str,
    row_limit: int,
) -> SnapshotCompareResponse:
    datasource = session.get(DataSource, datasource_id)

    if not datasource:
        raise DataSourceNotFoundError(datasource_id)
    if datasource.source_type != DataSourceType.ICEBERG:
        raise DataSourceValidationError(
            'Snapshot comparison is only available for Iceberg datasources',
            details={'datasource_id': datasource_id},
        )

    config_base = {
        'source_type': datasource.source_type,
        **datasource.config,
    }
    config_a = {**config_base, 'snapshot_id': snapshot_a}
    config_b = {**config_base, 'snapshot_id': snapshot_b}

    lf_a = load_datasource(config_a)
    lf_b = load_datasource(config_b)

    schema_a = lf_a.collect_schema()
    schema_b = lf_b.collect_schema()

    row_count_a = lf_a.select(pl.len()).collect().item()
    row_count_b = lf_b.select(pl.len()).collect().item()

    stats_a = _build_snapshot_stats(lf_a, schema_a)
    stats_b = _build_snapshot_stats(lf_b, schema_b)

    preview_a = _build_snapshot_preview(lf_a, schema_a, row_limit)
    preview_b = _build_snapshot_preview(lf_b, schema_b, row_limit)

    diff = _build_schema_diff(schema_a, schema_b)

    return SnapshotCompareResponse(
        datasource_id=datasource_id,
        snapshot_a=snapshot_a,
        snapshot_b=snapshot_b,
        row_count_a=row_count_a,
        row_count_b=row_count_b,
        row_count_delta=row_count_b - row_count_a,
        schema_diff=diff,
        stats_a=stats_a,
        stats_b=stats_b,
        preview_a=preview_a,
        preview_b=preview_b,
    )


def _build_snapshot_preview(
    lazy: pl.LazyFrame,
    schema: pl.Schema,
    row_limit: int,
) -> SnapshotPreview:
    data = lazy.limit(row_limit).collect().to_dicts()
    return SnapshotPreview(
        columns=list(schema.keys()),
        column_types={name: str(dtype) for name, dtype in schema.items()},
        data=data,
        row_count=len(data),
    )


def _supports_min_max(dtype: pl.DataType) -> bool:
    return isinstance(
        dtype,
        (
            pl.Int8,
            pl.Int16,
            pl.Int32,
            pl.Int64,
            pl.UInt8,
            pl.UInt16,
            pl.UInt32,
            pl.UInt64,
            pl.Float32,
            pl.Float64,
            pl.Utf8,
            pl.Date,
            pl.Datetime,
            pl.Time,
        ),
    )


def _supports_unique(dtype: pl.DataType) -> bool:
    return isinstance(
        dtype,
        (
            pl.Int8,
            pl.Int16,
            pl.Int32,
            pl.Int64,
            pl.UInt8,
            pl.UInt16,
            pl.UInt32,
            pl.UInt64,
            pl.Float32,
            pl.Float64,
            pl.Utf8,
            pl.Boolean,
            pl.Date,
            pl.Datetime,
            pl.Time,
        ),
    )


def _build_snapshot_stats(lazy: pl.LazyFrame, schema: pl.Schema) -> list[ColumnStats]:
    exprs: list[pl.Expr] = []
    for name, dtype in schema.items():
        exprs.append(pl.col(name).null_count().alias(f'{name}__null_count'))
        if _supports_unique(dtype):
            exprs.append(pl.col(name).drop_nulls().n_unique().alias(f'{name}__unique_count'))
        if _supports_min_max(dtype):
            exprs.append(pl.col(name).min().alias(f'{name}__min'))
            exprs.append(pl.col(name).max().alias(f'{name}__max'))

    stats_frame = lazy.select(exprs).collect()
    results: list[ColumnStats] = []
    for name, dtype in schema.items():
        null_count = int(stats_frame[f'{name}__null_count'][0])
        unique_count = None
        if f'{name}__unique_count' in stats_frame.columns:
            unique_count = int(stats_frame[f'{name}__unique_count'][0])
        min_val = None
        max_val = None
        if f'{name}__min' in stats_frame.columns:
            min_val = stats_frame[f'{name}__min'][0]
        if f'{name}__max' in stats_frame.columns:
            max_val = stats_frame[f'{name}__max'][0]
        results.append(
            ColumnStats(
                column=name,
                dtype=str(dtype),
                null_count=null_count,
                unique_count=unique_count,
                min=min_val,
                max=max_val,
            )
        )
    return results


def _build_schema_diff(schema_a: pl.Schema, schema_b: pl.Schema) -> list[SchemaDiff]:
    diffs: list[SchemaDiff] = []
    cols_a = set(schema_a.keys())
    cols_b = set(schema_b.keys())

    for name in sorted(cols_a - cols_b):
        diffs.append(
            SchemaDiff(
                column=name,
                status='removed',
                type_a=str(schema_a[name]),
                type_b=None,
            )
        )

    for name in sorted(cols_b - cols_a):
        diffs.append(
            SchemaDiff(
                column=name,
                status='added',
                type_a=None,
                type_b=str(schema_b[name]),
            )
        )

    for name in sorted(cols_a & cols_b):
        dtype_a = str(schema_a[name])
        dtype_b = str(schema_b[name])
        if dtype_a == dtype_b:
            continue
        diffs.append(
            SchemaDiff(
                column=name,
                status='type_changed',
                type_a=dtype_a,
                type_b=dtype_b,
            )
        )

    return diffs


def get_datasource(session: Session, datasource_id: str) -> DataSourceResponse:
    datasource = session.get(DataSource, datasource_id)

    if not datasource:
        raise DataSourceNotFoundError(datasource_id)

    response = DataSourceResponse.model_validate(datasource)
    response.output_of_tab_id = datasource.config.get('analysis_tab_id') if isinstance(datasource.config, dict) else None
    return response


def list_datasources(session: Session, include_hidden: bool = False) -> list[DataSourceResponse]:
    query = select(DataSource)
    if not include_hidden:
        # SQLModel field typed as bool; == creates SA expression at runtime
        query = query.where(col(DataSource.is_hidden) == False)  # type: ignore[arg-type]  # noqa: E712
    result = session.execute(query)
    datasources = result.scalars().all()

    # Populate schema_cache for datasources that don't have it
    for ds in datasources:
        if ds.schema_cache is not None:
            continue
        if ds.source_type == 'analysis' or ds.created_by == 'analysis':
            continue
        try:
            schema_info = _extract_schema(ds)
            schema_cache = schema_info.model_dump()
        except Exception as e:
            logger.warning(f'Failed to extract schema for datasource {ds.id}: {e}')
            continue
        session.execute(
            update(DataSource)
            .where(DataSource.id == ds.id)  # type: ignore[arg-type]
            .values(schema_cache=schema_cache)
        )
        session.commit()
        ds.schema_cache = schema_cache
        logger.info(f'Populated schema cache for datasource {ds.id}')
    results: list[DataSourceResponse] = []
    for ds in datasources:
        item = DataSourceResponse.model_validate(ds)
        item.output_of_tab_id = ds.config.get('analysis_tab_id') if isinstance(ds.config, dict) else None
        results.append(item)
    return results


def update_datasource(
    session: Session,
    datasource_id: str,
    update: DataSourceUpdate,
) -> DataSourceResponse:
    datasource = session.get(DataSource, datasource_id)

    if not datasource:
        raise DataSourceNotFoundError(datasource_id)

    # Update name if provided
    if update.name is not None:
        datasource.name = update.name

    # Update is_hidden if provided
    if update.is_hidden is not None:
        datasource.is_hidden = update.is_hidden

    # Update config if provided
    if update.config is not None:
        if 'column_schema' in update.config:
            raise DataSourceValidationError(
                'Datasource schemas are read-only and cannot be modified',
                details={'datasource_id': datasource_id},
            )

        immutable_keys = {
            'file': ['file_path'],
            'database': ['connection_string'],
            'duckdb': ['db_path'],
            'iceberg': ['metadata_path'],
        }
        for key in immutable_keys.get(datasource.source_type, []):
            if key not in update.config:
                continue
            if update.config.get(key) == datasource.config.get(key):
                continue
            raise DataSourceValidationError(
                'Datasource location is immutable. Create a new datasource to change location.',
                details={'datasource_id': datasource_id, 'field': key},
            )

        if (
            datasource.source_type == 'duckdb'
            and 'read_only' in update.config
            and update.config.get('read_only') != datasource.config.get('read_only', True)
        ):
            raise DataSourceValidationError(
                'Datasource mode is immutable. Create a new datasource to change read-only mode.',
                details={'datasource_id': datasource_id, 'field': 'read_only'},
            )

        # Check if parsing options changed (requires schema re-extraction)
        parsing_keys = [
            'csv_options',
            'sheet_name',
            'start_row',
            'start_col',
            'end_col',
            'end_row',
            'has_header',
            'skip_rows',
            'table_name',
            'named_range',
            'cell_range',
        ]
        parsing_changed = any(key in update.config for key in parsing_keys)

        next_config = {**datasource.config, **update.config}
        has_excel_bounds = any(
            key in update.config
            for key in ['sheet_name', 'start_row', 'start_col', 'end_col', 'end_row', 'table_name', 'named_range', 'cell_range']
        )
        is_excel_file = next_config.get('file_type') == 'excel'
        if datasource.source_type == 'file' and is_excel_file and has_excel_bounds:
            file_path = next_config.get('file_path')
            if not file_path:
                raise DataSourceValidationError(
                    'Excel datasource requires file_path',
                    details={'datasource_id': datasource_id},
                )
            start_row = next_config.get('start_row')
            if start_row is None:
                start_row = 0
            start_col = next_config.get('start_col')
            if start_col is None:
                start_col = 0
            end_col = next_config.get('end_col')
            if end_col is None:
                end_col = 0
            try:
                resolved_sheet, resolved_start_row, resolved_start_col, resolved_end_col, resolved_end_row = resolve_excel_selection(
                    Path(file_path),
                    next_config.get('sheet_name'),
                    int(start_row),
                    int(start_col),
                    int(end_col),
                    next_config.get('end_row'),
                    next_config.get('table_name'),
                    next_config.get('named_range'),
                    next_config.get('cell_range'),
                )
            except Exception as exc:
                raise DataSourceValidationError(
                    str(exc),
                    details={'datasource_id': datasource_id},
                ) from exc
            next_config = {
                **next_config,
                'sheet_name': resolved_sheet,
                'start_row': resolved_start_row,
                'start_col': resolved_start_col,
                'end_col': resolved_end_col,
                'end_row': resolved_end_row,
            }

        # Merge new config with existing config
        datasource.config = next_config
        if parsing_changed:
            datasource.schema_cache = None

    session.commit()
    session.refresh(datasource)

    # Only log engine run for non-metadata updates (is_hidden toggle is metadata-only)
    has_config_or_name_update = update.name is not None or update.config is not None
    if has_config_or_name_update:
        branch = datasource.config.get('branch', 'master') if isinstance(datasource.config, dict) else 'master'
        update_request = update.model_dump(exclude_none=True)
        update_request['iceberg_options'] = {'branch': branch}
        run_payload = engine_run_service.create_engine_run_payload(
            analysis_id=None,
            datasource_id=datasource_id,
            kind='datasource_update',
            status='success',
            request_json=update_request,
            result_json={'datasource_id': datasource_id, 'datasource_name': datasource.name},
        )
        engine_run_service.create_engine_run(session, run_payload)

    logger.info(f'Updated datasource {datasource_id}')
    response = DataSourceResponse.model_validate(datasource)
    response.output_of_tab_id = datasource.config.get('analysis_tab_id') if isinstance(datasource.config, dict) else None
    return response


def _compute_histogram(series: pl.Series, bins: int = 20) -> list[dict[str, object]]:
    if series.is_empty():
        return []
    stats = series.drop_nulls()
    if stats.is_empty():
        return []
    stats = stats.cast(pl.Float64, strict=False)
    min_raw: Any = stats.min()
    max_raw: Any = stats.max()
    if min_raw is None or max_raw is None:
        return []
    min_val = float(min_raw)
    max_val = float(max_raw)
    if min_val == max_val:
        return [{'start': min_val, 'end': max_val, 'count': stats.len()}]
    width = (max_val - min_val) / bins
    result: list[dict[str, object]] = []
    for i in range(bins):
        start = min_val + i * width
        end = min_val + (i + 1) * width
        bin_count = (
            series.filter((series >= start) & (series <= end)).len()
            if i == bins - 1
            else series.filter((series >= start) & (series < end)).len()
        )
        result.append({'start': round(start, 4), 'end': round(end, 4), 'count': bin_count})
    return result


def get_column_stats(
    session: Session,
    datasource_id: str,
    column_name: str,
    use_sample: bool = True,
    sample_size: int = 10000,
    datasource_config: dict | None = None,
) -> ColumnStatsResponse:
    datasource = session.get(DataSource, datasource_id)

    if not datasource:
        raise DataSourceNotFoundError(datasource_id)

    config = {
        'source_type': datasource.source_type,
        **datasource.config,
    }
    if datasource_config:
        config = {**config, **datasource_config}
    lazy = load_datasource(config)

    schema = lazy.collect_schema()
    if column_name not in schema:
        raise ValueError(f'Column not found: {column_name}')

    if use_sample:
        lazy = lazy.head(sample_size)  # type: ignore[attr-defined]

    frame = lazy.select([pl.col(column_name)]).collect()
    series = frame[column_name]
    dtype = schema[column_name]

    count = series.len()
    null_count = series.null_count()
    null_percentage = (null_count / count * 100.0) if count > 0 else 0.0

    stats: dict[str, object] = {
        'column': column_name,
        'dtype': str(dtype),
        'count': count,
        'null_count': null_count,
        'null_percentage': null_percentage,
    }

    if isinstance(dtype, (pl.Int8, pl.Int16, pl.Int32, pl.Int64, pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64, pl.Float32, pl.Float64)):
        non_null = series.drop_nulls()
        stats.update(
            {
                'mean': series.mean(),
                'std': series.std(),
                'min': series.min(),
                'max': series.max(),
                'median': series.median(),
                'q25': series.quantile(0.25),
                'q75': series.quantile(0.75),
                'histogram': _compute_histogram(non_null),
            }
        )
        return ColumnStatsResponse.model_validate(stats)

    if isinstance(dtype, pl.Utf8):
        length_series = series.str.len_chars()  # type: ignore[attr-defined]
        stats.update(
            {
                'unique': series.n_unique(),
                'min_length': length_series.min(),
                'max_length': length_series.max(),
                'avg_length': length_series.mean(),
                'top_values': (series.value_counts().sort('count', descending=True).head(5).to_dicts()),
            }
        )
        return ColumnStatsResponse.model_validate(stats)

    if isinstance(dtype, pl.Datetime):
        stats.update(
            {
                'min': series.min(),
                'max': series.max(),
            }
        )
        return ColumnStatsResponse.model_validate(stats)

    if isinstance(dtype, pl.Boolean):
        value_counts = series.value_counts().to_dicts()
        true_count = 0
        false_count = 0
        for item in value_counts:
            value = item.get(column_name)
            if value is True:
                true_count = int(item.get('count', 0))
            if value is False:
                false_count = int(item.get('count', 0))
        stats.update(
            {
                'true_count': true_count,
                'false_count': false_count,
            }
        )
        return ColumnStatsResponse.model_validate(stats)

    stats.update({'unique': series.n_unique()})
    return ColumnStatsResponse.model_validate(stats)


def delete_datasource(session: Session, datasource_id: str) -> None:
    datasource = session.get(DataSource, datasource_id)

    if not datasource:
        raise DataSourceNotFoundError(datasource_id)

    _delete_datasource_files(datasource)

    session.delete(datasource)
    session.commit()
    logger.info(f'Deleted datasource {datasource_id}')


def _delete_datasource_files(datasource: DataSource) -> None:
    # Delete associated file if it's a file datasource (original file in uploads)
    if datasource.source_type == 'file' and 'file_path' in datasource.config:
        file_path = Path(datasource.config['file_path'])
        if file_path.exists():
            try:
                # Check if file is accessible before deletion
                if not file_path.is_file():
                    logger.warning(f'Path exists but is not a file: {file_path}')
                else:
                    file_path.unlink()
                    logger.info(f'Deleted file: {file_path}')
            except PermissionError as e:
                logger.error(f'Permission denied when deleting file {file_path}: {e}')
                raise FileError(
                    f'Permission denied when deleting file: {file_path}',
                    error_code='FILE_PERMISSION_DENIED',
                    details={'file_path': str(file_path)},
                )
            except OSError as e:
                logger.error(f'OS error when deleting file {file_path}: {e}')
                raise FileError(
                    f'Failed to delete file: {file_path}',
                    error_code='FILE_DELETE_ERROR',
                    details={'file_path': str(file_path), 'error': str(e)},
                )

    # Delete Iceberg datasource files from clean or exports directory
    if datasource.source_type == 'iceberg' and isinstance(datasource.config, dict):
        config = datasource.config
        metadata_path = config.get('metadata_path')
        if metadata_path:
            metadata_path_obj = Path(metadata_path)
            paths = namespace_paths()
            exports_dir = paths.exports_dir
            clean_dir = paths.clean_dir

            # Determine if this is in exports_dir or clean_dir
            is_under_exports = exports_dir in metadata_path_obj.parents or metadata_path_obj.parent == exports_dir
            is_under_clean = clean_dir in metadata_path_obj.parents or metadata_path_obj.parent == clean_dir

            if is_under_exports or is_under_clean:
                # This is an exported or ingested datasource - delete the entire directory
                parent_dir = metadata_path_obj.parent
                try:
                    if parent_dir.exists() and parent_dir.is_dir():
                        shutil.rmtree(parent_dir)
                        logger.info(f'Deleted Iceberg directory: {parent_dir}')
                except OSError as e:
                    logger.error(f'OS error when deleting Iceberg directory {parent_dir}: {e}')
                    raise FileError(
                        f'Failed to delete Iceberg directory: {parent_dir}',
                        error_code='FILE_DELETE_ERROR',
                        details={'path': str(parent_dir), 'error': str(e)},
                    )
