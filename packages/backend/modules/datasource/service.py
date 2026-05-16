import logging
import os
import shutil
import uuid
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from contracts.datasource.models import DataSource, DataSourceColumnMetadata, DataSourceCreatedBy
from contracts.datasource.source_types import DataSourceType
from core.exceptions import (
    DataSourceNotFoundError,
    DataSourceValidationError,
    FileError,
)
from core.namespace import namespace_paths
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from sqlalchemy import select
from sqlalchemy.orm import defer
from sqlmodel import Session, col

from modules.datasource.schemas import (
    BatchColumnDescriptionUpdate,
    ColumnDescriptionPatch,
    DataSourceDescriptionModel,
    DataSourceListItem,
    DataSourceResponse,
    DataSourceUpdate,
    FileListItem,
    FileListResponse,
    SchemaInfo,
)

logger = logging.getLogger(__name__)


def _open_excel_workbook(file_path: Path, *, table_name: str | None) -> Any:
    # Table metadata is unavailable in openpyxl read_only mode.
    return load_workbook(file_path, read_only=table_name is None, data_only=True)


def create_placeholder_output_datasource(
    session: Session,
    result_id: str,
    analysis_id: str,
    analysis_tab_id: str,
) -> None:
    try:
        uuid.UUID(result_id)
    except ValueError:
        raise ValueError(f"result_id must be a valid UUID, got: {result_id!r}") from None
    existing = session.get(DataSource, result_id)
    if existing:
        existing_owner = existing.created_by_analysis_id
        if existing_owner is not None and str(existing_owner) != analysis_id:
            raise ValueError(
                f"Output result_id '{result_id}' is already owned by analysis '{existing_owner}', cannot reuse it in analysis '{analysis_id}'",
            )
        if existing_owner is None and existing.created_by != DataSourceCreatedBy.ANALYSIS.value:
            raise ValueError(
                f"Output result_id '{result_id}' conflicts with an existing datasource not managed by analysis outputs",
            )
        next_config = dict(existing.config) if isinstance(existing.config, dict) else {}
        next_config["analysis_tab_id"] = analysis_tab_id
        existing.config = next_config
        existing.created_by_analysis_id = analysis_id
        existing.created_by = DataSourceCreatedBy.ANALYSIS.value
        session.add(existing)
        session.flush()
        return
    datasource = DataSource(
        id=result_id,
        name=result_id,
        source_type=DataSourceType.ANALYSIS,
        config={"analysis_tab_id": analysis_tab_id},
        created_by_analysis_id=analysis_id,
        created_by=DataSourceCreatedBy.ANALYSIS.value,
        is_hidden=True,
        created_at=datetime.now(UTC).replace(tzinfo=None),
    )
    session.add(datasource)
    session.flush()


def create_analysis_datasource(
    session: Session,
    name: str,
    description: str | None,
    analysis_id: str,
    analysis_tab_id: str | None = None,
    is_hidden: bool = False,
    source_type: DataSourceType = DataSourceType.ANALYSIS,
) -> DataSourceResponse:
    from contracts.analysis.models import Analysis

    analysis = session.get(Analysis, analysis_id)
    if not analysis:
        raise ValueError(f"Analysis {analysis_id} not found")
    datasource_id = str(uuid.uuid4())
    config = {}
    if analysis_tab_id:
        config["analysis_tab_id"] = analysis_tab_id

    datasource = DataSource(
        id=datasource_id,
        name=name,
        description=DataSourceDescriptionModel.normalize_description(description),
        source_type=source_type,
        config=config,
        created_by_analysis_id=analysis_id,
        created_by=DataSourceCreatedBy.ANALYSIS.value,
        is_hidden=is_hidden,
        created_at=datetime.now(UTC).replace(tzinfo=None),
    )

    session.add(datasource)
    session.commit()
    session.refresh(datasource)

    return DataSourceResponse.model_validate(datasource)


@dataclass
class ExcelPreviewResult:
    preview: list[list[str | None]]
    detected_end_row: int | None
    sheet_name: str
    start_row: int
    start_col: int
    end_col: int
    end_row: int


def build_excel_preview(
    file_path: Path,
    sheet_name: str,
    start_row: int,
    start_col: int,
    end_col: int,
    end_row: int | None,
    has_header: bool,
    table_name: str | None = None,
    named_range: str | None = None,
    cell_range: str | None = None,
    preview_rows: int = 100,
) -> ExcelPreviewResult:
    workbook = _open_excel_workbook(file_path, table_name=table_name)
    try:
        resolved = _resolve_excel_bounds(
            workbook,
            sheet_name,
            start_row,
            start_col,
            end_col,
            end_row,
            table_name,
            named_range,
            cell_range,
        )
        sheet = workbook[resolved.sheet_name]

        end_row_value = resolved.end_row
        if end_row_value is None:
            end_row_value = _detect_end_row(sheet, resolved.start_row, resolved.start_col, resolved.end_col)
        _validate_excel_bounds(
            sheet,
            resolved.start_row,
            resolved.start_col,
            resolved.end_col,
            end_row_value,
        )

        preview_end_row = min(resolved.start_row + preview_rows - 1, end_row_value)
        rows = _collect_preview_rows(
            sheet,
            resolved.start_row,
            resolved.start_col,
            resolved.end_col,
            preview_end_row,
        )
        return ExcelPreviewResult(
            preview=rows,
            detected_end_row=end_row_value,
            sheet_name=resolved.sheet_name,
            start_row=resolved.start_row,
            start_col=resolved.start_col,
            end_col=resolved.end_col,
            end_row=end_row_value,
        )
    finally:
        workbook.close()


def resolve_excel_selection(
    file_path: Path,
    sheet_name: str | None,
    start_row: int,
    start_col: int,
    end_col: int,
    end_row: int | None,
    table_name: str | None = None,
    named_range: str | None = None,
    cell_range: str | None = None,
) -> tuple[str, int, int, int, int]:
    try:
        workbook = _open_excel_workbook(file_path, table_name=table_name)
        try:
            target_sheet = sheet_name or (workbook.sheetnames[0] if workbook.sheetnames else None)
            if not target_sheet:
                raise ValueError("No sheets found in file")
            resolved = _resolve_excel_bounds(
                workbook,
                target_sheet,
                start_row,
                start_col,
                end_col,
                end_row,
                table_name,
                named_range,
                cell_range,
            )
            sheet = workbook[resolved.sheet_name]
            end_row_value = resolved.end_row
            if end_row_value is None:
                end_row_value = _detect_end_row(sheet, resolved.start_row, resolved.start_col, resolved.end_col)
            _validate_excel_bounds(
                sheet,
                resolved.start_row,
                resolved.start_col,
                resolved.end_col,
                end_row_value,
            )
            return (
                resolved.sheet_name,
                resolved.start_row,
                resolved.start_col,
                resolved.end_col,
                end_row_value,
            )
        finally:
            workbook.close()
    except ValueError as exc:
        raise DataSourceValidationError(str(exc), details={"file_path": str(file_path)}) from exc


@dataclass
class _ExcelBounds:
    sheet_name: str
    start_row: int
    start_col: int
    end_col: int
    end_row: int | None


def _resolve_excel_bounds(
    workbook,
    sheet_name: str,
    start_row: int,
    start_col: int,
    end_col: int,
    end_row: int | None,
    table_name: str | None,
    named_range: str | None,
    cell_range: str | None,
) -> _ExcelBounds:
    if table_name:
        sheet = workbook[sheet_name]
        tables = getattr(sheet, "tables", None)
        if not tables:
            raise ValueError(f"No tables available in sheet: {sheet_name}")
        table = tables.get(table_name)
        if not table:
            raise ValueError(f"Table not found: {table_name}")
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if min_col is None or min_row is None or max_col is None or max_row is None:
            raise ValueError(f"Invalid table range: {table_name}")
        min_col = int(min_col)
        min_row = int(min_row)
        max_col = int(max_col)
        max_row = int(max_row)
        return _ExcelBounds(sheet_name, min_row - 1, min_col - 1, max_col - 1, max_row - 1)

    if named_range:
        defined = workbook.defined_names.get(named_range)
        if not defined:
            raise ValueError(f"Named range not found: {named_range}")
        destinations = list(defined.destinations)
        if not destinations:
            raise ValueError(f"Named range has no destinations: {named_range}")
        dest_sheet, coord = destinations[0]
        min_col, min_row, max_col, max_row = range_boundaries(coord)
        if min_col is None or min_row is None or max_col is None or max_row is None:
            raise ValueError(f"Invalid named range: {named_range}")
        min_col = int(min_col)
        min_row = int(min_row)
        max_col = int(max_col)
        max_row = int(max_row)
        return _ExcelBounds(dest_sheet, min_row - 1, min_col - 1, max_col - 1, max_row - 1)

    if cell_range:
        return _parse_cell_range(workbook, cell_range, sheet_name)

    resolved_start_row = max(start_row, 0)
    resolved_start_col = max(start_col, 0)
    resolved_end_col = end_col
    if resolved_end_col <= resolved_start_col:
        sheet = workbook[sheet_name]
        resolved_end_col = _detect_end_col(sheet, resolved_start_row, resolved_start_col)
    resolved_end_col = max(resolved_end_col, resolved_start_col)
    end_row_value = end_row
    if end_row_value is not None:
        end_row_value = max(end_row_value, resolved_start_row)
    return _ExcelBounds(
        sheet_name,
        resolved_start_row,
        resolved_start_col,
        resolved_end_col,
        end_row_value,
    )


def _parse_cell_range(workbook, cell_range: str, default_sheet: str | None) -> _ExcelBounds:
    raw = cell_range.strip()
    if not raw:
        raise ValueError("Cell range cannot be empty")
    target_sheet = default_sheet
    coord = raw
    if "!" in raw:
        sheet_part, coord_part = raw.split("!", maxsplit=1)
        sheet_part = sheet_part.strip()
        if sheet_part.startswith("'") and sheet_part.endswith("'"):
            sheet_part = sheet_part[1:-1]
        if not sheet_part:
            raise ValueError(f"Invalid cell range sheet: {cell_range}")
        target_sheet = sheet_part
        coord = coord_part.strip()
    if not target_sheet:
        target_sheet = workbook.sheetnames[0] if workbook.sheetnames else None
    if not target_sheet or target_sheet not in workbook.sheetnames:
        raise ValueError(f"Sheet not found for cell range: {target_sheet}")
    min_col, min_row, max_col, max_row = range_boundaries(coord)
    if min_col is None or min_row is None or max_col is None or max_row is None:
        raise ValueError(f"Invalid cell range: {cell_range}")
    return _ExcelBounds(
        target_sheet,
        int(min_row) - 1,
        int(min_col) - 1,
        int(max_col) - 1,
        int(max_row) - 1,
    )


def format_excel_cell_range(
    sheet_name: str,
    start_row: int,
    start_col: int,
    end_row: int,
    end_col: int,
) -> str:
    start_cell = f"{get_column_letter(start_col + 1)}{start_row + 1}"
    end_cell = f"{get_column_letter(end_col + 1)}{end_row + 1}"
    return f"{sheet_name}!{start_cell}:{end_cell}"


def _validate_excel_bounds(sheet, start_row: int, start_col: int, end_col: int, end_row: int) -> None:
    if start_row < 0 or start_col < 0:
        raise ValueError("Excel bounds must be non-negative")
    if end_row < start_row or end_col < start_col:
        raise ValueError("Excel bounds are invalid")
    max_row = sheet.max_row or 0
    max_col = sheet.max_column or 0
    if max_row <= 0 or max_col <= 0:
        raise ValueError("Excel sheet has no data")
    if start_row >= max_row or end_row >= max_row:
        raise ValueError("Excel row bounds exceed sheet size")
    if start_col >= max_col or end_col >= max_col:
        raise ValueError("Excel column bounds exceed sheet size")


def _detect_end_col(sheet, start_row: int, start_col: int) -> int:
    """Scan the header row (start_row) rightward to find the last non-empty column."""
    max_col = sheet.max_column or 0
    if max_col <= start_col:
        return start_col
    last_col = start_col
    for cell in sheet.iter_rows(
        min_row=start_row + 1,
        max_row=start_row + 1,
        min_col=start_col + 1,
        max_col=max_col,
    ):
        for c in cell:
            if c.value is not None and str(c.value).strip() != "":
                last_col = c.column - 1  # 0-indexed
    return last_col


def _detect_end_row(sheet, start_row: int, start_col: int, end_col: int) -> int:
    max_row = sheet.max_row or 0
    if max_row <= start_row:
        return start_row
    for row_index in range(start_row + 1, max_row + 1):
        values: list[object | None] = []
        for cell in sheet.iter_rows(
            min_row=row_index,
            max_row=row_index,
            min_col=start_col + 1,
            max_col=end_col + 1,
        ):
            values = [c.value for c in cell]
        if all(value is None or str(value).strip() == "" for value in values):
            return max(start_row, row_index - 2)
    return max_row - 1


def _collect_preview_rows(
    sheet,
    start_row: int,
    start_col: int,
    end_col: int,
    preview_end_row: int,
) -> list[list[str | None]]:
    rows: list[list[str | None]] = []
    for row in sheet.iter_rows(
        min_row=start_row + 1,
        max_row=preview_end_row + 1,
        min_col=start_col + 1,
        max_col=end_col + 1,
    ):
        values = [cell.value for cell in row]
        rows.append([str(value) if value is not None else None for value in values])
    return rows


def _get_column_metadata_map(session: Session, datasource_id: str) -> dict[str, str | None]:
    rows = session.execute(
        select(DataSourceColumnMetadata).where(DataSourceColumnMetadata.datasource_id == datasource_id),  # type: ignore[arg-type]
    ).scalars()
    return {row.column_name: row.description for row in rows}


def attach_column_descriptions(
    session: Session,
    datasource_id: str,
    schema_info: SchemaInfo,
) -> SchemaInfo:
    descriptions = _get_column_metadata_map(session, datasource_id)
    columns = [col.model_copy(update={"description": descriptions.get(col.name)}) for col in schema_info.columns]
    return schema_info.model_copy(update={"columns": columns})


def update_column_descriptions(
    session: Session,
    datasource_id: str,
    payload: BatchColumnDescriptionUpdate,
    schema_info: SchemaInfo,
) -> SchemaInfo:
    datasource = session.get(DataSource, datasource_id)
    if not datasource:
        raise DataSourceNotFoundError(datasource_id)

    active_columns = {column.name for column in schema_info.columns}

    for patch in payload.columns:
        if patch.column_name not in active_columns:
            raise DataSourceValidationError(
                f"Column not found in active schema: {patch.column_name}",
                details={
                    "datasource_id": datasource_id,
                    "column_name": patch.column_name,
                },
            )

    existing = session.execute(
        select(DataSourceColumnMetadata).where(DataSourceColumnMetadata.datasource_id == datasource_id),  # type: ignore[arg-type]
    ).scalars()
    existing_by_name = {row.column_name: row for row in existing}
    now = datetime.now(UTC).replace(tzinfo=None)

    for patch in payload.columns:
        description = ColumnDescriptionPatch.normalize_description(patch.description)
        row = existing_by_name.get(patch.column_name)
        if description is None:
            if row is not None:
                session.delete(row)
            continue
        if row is None:
            session.add(
                DataSourceColumnMetadata(
                    id=str(uuid.uuid4()),
                    datasource_id=datasource_id,
                    column_name=patch.column_name,
                    description=description,
                    created_at=now,
                    updated_at=now,
                ),
            )
            continue
        row.description = description
        row.updated_at = now
        session.add(row)

    session.commit()
    return attach_column_descriptions(session, datasource_id, schema_info)


def list_data_files(path: str | None) -> FileListResponse:
    base_dir = Path(os.path.realpath(namespace_paths().base_dir))
    target = Path(path) if path else base_dir
    resolved = Path(os.path.realpath(target))
    if base_dir not in resolved.parents and base_dir != resolved:
        raise ValueError(f"Path must be inside data directory: {base_dir}")
    if not resolved.exists():
        raise ValueError(f"Path does not exist: {resolved}")
    if not resolved.is_dir():
        raise ValueError(f"Path must be a directory: {resolved}")

    entries = [
        FileListItem(
            name=item.name,
            path=str(item),
            is_dir=item.is_dir(),
        )
        for item in sorted(
            resolved.iterdir(),
            key=lambda entry: (not entry.is_dir(), entry.name.lower()),
        )
    ]
    return FileListResponse(base_path=str(resolved), entries=entries)


def get_datasource(session: Session, datasource_id: str) -> DataSourceResponse:
    datasource = session.get(DataSource, datasource_id)

    if not datasource:
        raise DataSourceNotFoundError(datasource_id)

    response = DataSourceResponse.model_validate(datasource)
    response.output_of_tab_id = datasource.config.get("analysis_tab_id") if isinstance(datasource.config, dict) else None
    return response


def list_datasources(session: Session, include_hidden: bool = False) -> list[DataSourceListItem]:
    query = select(DataSource).options(
        defer(DataSource.schema_cache),  # type: ignore[arg-type]
    )
    if not include_hidden:
        # SQLModel field typed as bool; == creates SA expression at runtime
        query = query.where(col(DataSource.is_hidden) == False)  # type: ignore[arg-type]  # noqa: E712
    datasources = session.execute(query).scalars().all()
    results: list[DataSourceListItem] = []
    for ds in datasources:
        item = DataSourceListItem.model_validate(ds)
        item.output_of_tab_id = ds.config.get("analysis_tab_id") if isinstance(ds.config, dict) else None
        results.append(item)
    return results


def update_datasource(
    session: Session,
    datasource_id: str,
    update: DataSourceUpdate,
) -> DataSourceResponse:
    datasource = session.get(DataSource, datasource_id)

    if not datasource:
        raise DataSourceNotFoundError(datasource_id)

    # Update name if provided
    if update.name is not None:
        datasource.name = update.name

    if "description" in update.model_fields_set:
        datasource.description = DataSourceDescriptionModel.normalize_description(update.description)

    # Update is_hidden if provided
    if update.is_hidden is not None:
        datasource.is_hidden = update.is_hidden

    # Update config if provided
    if update.config is not None:
        if "column_schema" in update.config:
            raise DataSourceValidationError(
                "Datasource schemas are read-only and cannot be modified",
                details={"datasource_id": datasource_id},
            )

        protected_snapshot_keys = {
            "snapshot_id",
            "snapshot_timestamp_ms",
            "current_snapshot_id",
            "current_snapshot_timestamp_ms",
            "time_travel_snapshot_id",
            "time_travel_snapshot_timestamp_ms",
            "time_travel_ui",
        }
        for key in protected_snapshot_keys:
            if key not in update.config:
                continue
            raise DataSourceValidationError(
                "Snapshot metadata fields are system-managed and cannot be modified",
                details={"datasource_id": datasource_id, "field": key},
            )

        source_type = datasource.source_type_kind()
        immutable_keys = {
            DataSourceType.FILE: ["file_path"],
            DataSourceType.DATABASE: ["connection_string"],
            DataSourceType.ICEBERG: ["metadata_path"],
        }
        for key in immutable_keys.get(source_type, []):
            if key not in update.config:
                continue
            if update.config.get(key) == datasource.config.get(key):
                continue
            raise DataSourceValidationError(
                "Datasource location is immutable. Create a new datasource to change location.",
                details={"datasource_id": datasource_id, "field": key},
            )

        # Check if parsing options changed (requires schema re-extraction)
        parsing_keys = [
            "csv_options",
            "sheet_name",
            "start_row",
            "start_col",
            "end_col",
            "end_row",
            "has_header",
            "skip_rows",
            "table_name",
            "named_range",
            "cell_range",
        ]
        parsing_changed = any(key in update.config for key in parsing_keys)

        next_config = {**datasource.config, **update.config}
        has_excel_bounds = any(
            key in update.config
            for key in [
                "sheet_name",
                "start_row",
                "start_col",
                "end_col",
                "end_row",
                "table_name",
                "named_range",
                "cell_range",
            ]
        )
        is_excel_file = next_config.get("file_type") == "excel"
        if source_type == DataSourceType.FILE and is_excel_file and has_excel_bounds:
            file_path = next_config.get("file_path")
            if not file_path:
                raise DataSourceValidationError(
                    "Excel datasource requires file_path",
                    details={"datasource_id": datasource_id},
                )
            start_row = next_config.get("start_row")
            if start_row is None:
                start_row = 0
            start_col = next_config.get("start_col")
            if start_col is None:
                start_col = 0
            end_col = next_config.get("end_col")
            if end_col is None:
                end_col = 0
            try:
                (
                    resolved_sheet,
                    resolved_start_row,
                    resolved_start_col,
                    resolved_end_col,
                    resolved_end_row,
                ) = resolve_excel_selection(
                    Path(file_path),
                    next_config.get("sheet_name"),
                    int(start_row),
                    int(start_col),
                    int(end_col),
                    next_config.get("end_row"),
                    next_config.get("table_name"),
                    next_config.get("named_range"),
                    next_config.get("cell_range"),
                )
            except Exception as exc:
                raise DataSourceValidationError(
                    str(exc),
                    details={"datasource_id": datasource_id},
                ) from exc
            next_config = {
                **next_config,
                "sheet_name": resolved_sheet,
                "start_row": resolved_start_row,
                "start_col": resolved_start_col,
                "end_col": resolved_end_col,
                "end_row": resolved_end_row,
            }

        # Merge new config with existing config
        datasource.config = next_config
        if parsing_changed:
            datasource.schema_cache = None

    session.commit()
    session.refresh(datasource)

    logger.info(f"Updated datasource {datasource_id}")
    response = DataSourceResponse.model_validate(datasource)
    response.output_of_tab_id = datasource.config.get("analysis_tab_id") if isinstance(datasource.config, dict) else None
    return response


def delete_datasource(session: Session, datasource_id: str) -> None:
    datasource = session.get(DataSource, datasource_id)

    if not datasource:
        raise DataSourceNotFoundError(datasource_id)

    _delete_datasource_files(datasource)

    session.delete(datasource)
    session.commit()
    logger.info(f"Deleted datasource {datasource_id}")


def _delete_file_path(file_path: str) -> None:
    path = Path(file_path)
    if not path.exists():
        return
    try:
        if not path.is_file():
            logger.warning(f"Path exists but is not a file: {path}")
            return
        path.unlink()
        logger.info(f"Deleted file: {path}")
    except PermissionError as exc:
        logger.error(f"Permission denied when deleting file {path}: {exc}")
        raise FileError(
            f"Permission denied when deleting file: {path}",
            error_code="FILE_PERMISSION_DENIED",
            details={"file_path": str(path)},
        ) from exc
    except OSError as exc:
        logger.error(f"OS error when deleting file {path}: {exc}")
        raise FileError(
            f"Failed to delete file: {path}",
            error_code="FILE_DELETE_ERROR",
            details={"file_path": str(path), "error": str(exc)},
        ) from exc


def _iceberg_cleanup_root(metadata_path: str) -> Path | None:
    path = Path(os.path.realpath(metadata_path))
    start = path if path.is_dir() else path.parent
    paths = namespace_paths()
    clean_dir = Path(os.path.realpath(paths.clean_dir))
    exports_dir = Path(os.path.realpath(paths.exports_dir))

    for candidate in [start, *start.parents]:
        if candidate.parent == clean_dir or candidate.parent == exports_dir:
            return candidate
        if candidate in (clean_dir, exports_dir):
            return None
    return None


def _is_within(path: Path, root: Path) -> bool:
    resolved_path = Path(os.path.realpath(path))
    resolved_root = Path(os.path.realpath(root))
    return resolved_root in resolved_path.parents or resolved_root == resolved_path


def _delete_datasource_files(datasource: DataSource) -> None:
    if datasource.source_type == DataSourceType.FILE and "file_path" in datasource.config:
        _delete_file_path(str(datasource.config["file_path"]))

    if datasource.source_type == DataSourceType.ICEBERG and isinstance(datasource.config, dict):
        config = datasource.config
        metadata_path = config.get("metadata_path")
        if isinstance(metadata_path, str):
            root = _iceberg_cleanup_root(metadata_path)
            if root:
                try:
                    if root.exists() and root.is_dir():
                        shutil.rmtree(root)
                        logger.info(f"Deleted Iceberg directory: {root}")
                except OSError as exc:
                    logger.error(f"OS error when deleting Iceberg directory {root}: {exc}")
                    raise FileError(
                        f"Failed to delete Iceberg directory: {root}",
                        error_code="FILE_DELETE_ERROR",
                        details={"path": str(root), "error": str(exc)},
                    ) from exc

        source = config.get("source")
        if not isinstance(source, dict):
            return
        if source.get("source_type") != DataSourceType.FILE:
            return
        file_path = source.get("file_path")
        if not isinstance(file_path, str):
            return
        if not _is_within(Path(file_path), namespace_paths().upload_dir):
            return
        _delete_file_path(file_path)
