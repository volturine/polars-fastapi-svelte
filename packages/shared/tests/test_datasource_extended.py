import asyncio
import threading
import uuid
from datetime import UTC, datetime, timedelta
from pathlib import Path
from unittest.mock import patch

import polars as pl
import pytest
from modules.datasource.service import _compute_histogram, create_analysis_datasource
from openpyxl import Workbook
from sqlmodel import select

from contracts.analysis.models import Analysis
from contracts.datasource.models import DataSource
from contracts.engine_runs.models import EngineRun
from core.exceptions import DataSourceValidationError


class TestDataSourceValidation:
    def test_upload_empty_file(self, client, temp_upload_dir: Path):
        files = {'file': ('empty.csv', b'', 'text/csv')}

        response = client.post('/api/v1/datasource/upload', files=files)

        # Should fail with validation error
        assert response.status_code in [400, 422]

    def test_upload_file_too_large(self, client, monkeypatch):
        # Create a large file (> 100MB)
        large_content = b'a' * (101 * 1024 * 1024)
        files = {'file': ('large.csv', large_content, 'text/csv')}

        response = client.post('/api/v1/datasource/upload', files=files)

        # Should fail with file too large error
        assert response.status_code in [400, 413, 422]

    def test_upload_unsupported_format(self, client):
        files = {'file': ('test.xyz', b'random data', 'application/octet-stream')}
        data = {'name': 'Unsupported Format Test'}

        response = client.post('/api/v1/datasource/upload', files=files, data=data)

        # Should fail with unsupported format error
        assert response.status_code in [400, 422]

    def test_upload_corrupted_csv(self, client):
        corrupted_csv = b'id,name,age\n1,Alice\n2,Bob,30,extra\n'
        files = {'file': ('corrupted.csv', corrupted_csv, 'text/csv')}
        data = {'name': 'Corrupted CSV Test'}

        response = client.post('/api/v1/datasource/upload', files=files, data=data)

        # May succeed but should handle gracefully
        if response.status_code == 200:
            json_data = response.json()
            assert 'id' in json_data

    def test_upload_csv_with_special_characters(self, client):
        csv_content = b'id,name,description\n1,"O\'Brien","Quote: \\"test\\""\n2,Smith,"Newline:\ntest"\n'
        files = {'file': ('special.csv', csv_content, 'text/csv')}
        data = {'name': 'Special Characters Test'}

        response = client.post('/api/v1/datasource/upload', files=files, data=data)

        assert response.status_code == 200

    def test_upload_csv_with_unicode(self, client):
        csv_content = 'id,name,city\n1,José,São Paulo\n2,François,Zürich\n'.encode()
        files = {'file': ('unicode.csv', csv_content, 'text/csv')}
        data = {'name': 'Unicode Test'}

        response = client.post('/api/v1/datasource/upload', files=files, data=data)

        assert response.status_code == 200
        json_data = response.json()
        assert 'id' in json_data

    def test_upload_with_missing_file(self, client):
        data = {'name': 'Missing File Test'}
        response = client.post('/api/v1/datasource/upload', data=data)

        assert response.status_code == 422

    def test_get_nonexistent_datasource(self, client):
        fake_id = str(uuid.uuid4())
        response = client.get(f'/api/v1/datasource/{fake_id}')

        assert response.status_code == 404

    def test_delete_nonexistent_datasource(self, client):
        fake_id = str(uuid.uuid4())
        response = client.delete(f'/api/v1/datasource/{fake_id}')

        assert response.status_code == 404

    def test_get_schema_nonexistent_datasource(self, client):
        fake_id = str(uuid.uuid4())
        response = client.get(f'/api/v1/datasource/{fake_id}/schema')

        assert response.status_code == 404

    def test_upload_csv_with_different_delimiters(self, client):
        # Tab-separated
        tsv_content = b'id\tname\tage\n1\tAlice\t25\n2\tBob\t30\n'
        files = {'file': ('test.tsv', tsv_content, 'text/tab-separated-values')}
        data = {'name': 'TSV Test'}

        response = client.post('/api/v1/datasource/upload', files=files, data=data)

        # Should handle or reject gracefully
        assert response.status_code in [200, 400, 422]

    def test_upload_json_array(self, client):
        """Test uploading JSON array format."""
        json_content = b'[{"id": 1, "name": "Alice"}, {"id": 2, "name": "Bob"}]'
        files = {'file': ('test.json', json_content, 'application/json')}
        data = {'name': 'JSON Array Test'}

        response = client.post('/api/v1/datasource/upload', files=files, data=data)

        assert response.status_code == 200

    def test_upload_ndjson_multiline(self, client):
        """Test uploading NDJSON with multiple lines."""
        ndjson_content = b'{"id": 1, "name": "Alice"}\n{"id": 2, "name": "Bob"}\n{"id": 3, "name": "Charlie"}\n'
        files = {'file': ('test.ndjson', ndjson_content, 'application/x-ndjson')}
        data = {'name': 'NDJSON Test'}

        response = client.post('/api/v1/datasource/upload', files=files, data=data)

        assert response.status_code == 200

    def test_upload_excel_xlsx(self, client, temp_upload_dir: Path):
        """Test uploading Excel XLSX file."""
        # Create a simple Excel file using Polars
        excel_path = temp_upload_dir / 'test.xlsx'
        df = pl.DataFrame({'id': [1, 2, 3], 'name': ['A', 'B', 'C']})

        df.write_excel(excel_path)

        with open(excel_path, 'rb') as f:
            files = {'file': ('test.xlsx', f.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
            data = {'name': 'Excel Test'}

        response = client.post('/api/v1/datasource/upload', files=files, data=data)

        assert response.status_code == 200

    def test_preflight_requires_xlsx(self, client):
        """Test preflight endpoint rejects non-xlsx files."""
        files = {'file': ('test.csv', b'a,b\n1,2', 'text/csv')}

        response = client.post('/api/v1/datasource/preflight', files=files)

        assert response.status_code == 400

    def test_preflight_excel_cell_range(self, client, temp_upload_dir: Path):
        """Test Excel preflight supports A1 range selection."""
        excel_path = temp_upload_dir / 'range.xlsx'
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            pytest.skip('Excel support not available')
        assert sheet is not None
        sheet.title = 'Sheet1'
        sheet.append(['id', 'name'])
        sheet.append([1, 'A'])
        sheet.append([2, 'B'])
        workbook.save(excel_path)

        with open(excel_path, 'rb') as f:
            files = {'file': ('range.xlsx', f.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'cell_range': 'A1:B3', 'has_header': 'true'}

        response = client.post('/api/v1/datasource/preflight', files=files, data=data)

        assert response.status_code == 200
        payload = response.json()
        assert payload['sheet_name'] == 'Sheet1'
        assert payload['start_row'] == 0
        assert payload['start_col'] == 0
        assert payload['end_col'] == 1
        assert payload['detected_end_row'] == 2
        assert len(payload['preview']) == 3

    def test_preflight_excel_invalid_cell_range(self, client, temp_upload_dir: Path):
        """Test Excel preflight rejects invalid cell ranges."""
        excel_path = temp_upload_dir / 'invalid.xlsx'
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            pytest.skip('Excel support not available')
        sheet.title = 'Sheet1'
        sheet.append(['id', 'name'])
        sheet.append([1, 'A'])
        workbook.save(excel_path)

        with open(excel_path, 'rb') as f:
            files = {'file': ('invalid.xlsx', f.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        data = {'cell_range': 'NotARange', 'has_header': 'true'}

        response = client.post('/api/v1/datasource/preflight', files=files, data=data)

        assert response.status_code == 400

    def test_preflight_excel_path_rejects_non_xlsx(self, client, temp_upload_dir: Path):
        csv_path = temp_upload_dir / 'invalid.csv'
        csv_path.write_text('a,b\n1,2')
        payload = {'file_path': str(csv_path)}

        response = client.post('/api/v1/datasource/preflight-path', json=payload)

        assert response.status_code == 400

    def test_preflight_excel_path_returns_preview(self, client, temp_upload_dir: Path, monkeypatch):
        excel_path = temp_upload_dir / 'path.xlsx'
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            pytest.skip('Excel support not available')
        sheet.title = 'Sheet1'
        sheet.append(['id', 'name'])
        sheet.append([1, 'A'])
        sheet.append([2, 'B'])
        workbook.save(excel_path)
        from modules.datasource import routes as datasource_routes

        from core import namespace

        def fake_paths():
            return namespace.NamespacePaths(
                base_dir=temp_upload_dir.parent,
                upload_dir=temp_upload_dir,
                clean_dir=temp_upload_dir.parent / 'clean',
                exports_dir=temp_upload_dir.parent / 'exports',
                db_path=temp_upload_dir.parent / 'namespace.db',
            )

        monkeypatch.setattr(namespace, 'namespace_paths', fake_paths)
        monkeypatch.setattr(datasource_routes, 'namespace_paths', fake_paths)
        payload = {'file_path': str(excel_path)}

        response = client.post('/api/v1/datasource/preflight-path', json=payload)

        assert response.status_code == 200
        body = response.json()
        assert body['sheet_name'] == 'Sheet1'
        assert body['preflight_id']
        assert len(body['preview']) == 3

    @pytest.mark.asyncio
    async def test_preflight_cleanup_removes_expired_without_clear_preflight(self, tmp_path: Path, monkeypatch) -> None:
        from modules.datasource import preflight

        keep_path = tmp_path / 'keep.xlsx'
        keep_path.write_bytes(b'keep')
        drop_path = tmp_path / 'drop.xlsx'
        drop_path.write_bytes(b'drop')
        stay_path = tmp_path / 'stay.xlsx'
        stay_path.write_bytes(b'stay')
        now = datetime.now(UTC).replace(tzinfo=None)

        async def fail(*args, **kwargs) -> None:
            raise AssertionError('_cleanup_expired() should not call clear_preflight()')

        monkeypatch.setattr(preflight, 'clear_preflight', fail)
        preflight._PREFLIGHTS.clear()
        preflight._PREFLIGHTS['keep'] = preflight.ExcelPreflight(
            temp_path=keep_path,
            sheets=[],
            tables={},
            named_ranges=[],
            created_at=now - preflight._PREFLIGHT_TTL - timedelta(seconds=1),
            delete_file=False,
        )
        preflight._PREFLIGHTS['drop'] = preflight.ExcelPreflight(
            temp_path=drop_path,
            sheets=[],
            tables={},
            named_ranges=[],
            created_at=now - preflight._PREFLIGHT_TTL - timedelta(seconds=1),
            delete_file=True,
        )
        preflight._PREFLIGHTS['stay'] = preflight.ExcelPreflight(
            temp_path=stay_path,
            sheets=[],
            tables={},
            named_ranges=[],
            created_at=now,
            delete_file=True,
        )

        try:
            await preflight._cleanup_expired()

            assert 'keep' not in preflight._PREFLIGHTS
            assert 'drop' not in preflight._PREFLIGHTS
            assert 'stay' in preflight._PREFLIGHTS
            assert keep_path.exists()
            assert not drop_path.exists()
            assert stay_path.exists()
        finally:
            preflight._PREFLIGHTS.clear()

    @pytest.mark.asyncio
    async def test_create_preflight_allows_concurrent_workbook_parsing(self, tmp_path: Path, monkeypatch) -> None:
        from modules.datasource import preflight

        class Sheet:
            def __init__(self, title: str) -> None:
                self.title = title
                self.tables = {f'{title}_table': object()}

        class Book:
            def __init__(self, title: str) -> None:
                self.sheetnames = [title]
                self.worksheets = [Sheet(title)]
                self.defined_names = [f'{title}_range']

        barrier = threading.Barrier(2, timeout=1)

        def fake_load_workbook(path: Path, read_only: bool = False, data_only: bool = True) -> Book:
            assert read_only is False
            assert data_only is True
            barrier.wait()
            return Book(path.stem)

        one = tmp_path / 'one.xlsx'
        two = tmp_path / 'two.xlsx'
        one.write_bytes(b'one')
        two.write_bytes(b'two')
        monkeypatch.setattr(preflight, 'load_workbook', fake_load_workbook)
        preflight._PREFLIGHTS.clear()

        try:
            first, second = await asyncio.wait_for(
                asyncio.gather(
                    preflight.create_preflight(one),
                    preflight.create_preflight(two),
                ),
                timeout=2,
            )
            first_id, first_preflight = first
            second_id, second_preflight = second

            assert first_id != second_id
            assert first_preflight.sheets == ['one']
            assert second_preflight.sheets == ['two']
            assert preflight._PREFLIGHTS[first_id] is first_preflight
            assert preflight._PREFLIGHTS[second_id] is second_preflight
        finally:
            preflight._PREFLIGHTS.clear()

    def test_confirm_excel_end_row(self, client, temp_upload_dir: Path):
        """Test Excel confirm stores manual end row selection."""
        excel_path = temp_upload_dir / 'bounds.xlsx'
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            pytest.skip('Excel support not available')
        assert sheet is not None
        sheet.title = 'Sheet1'
        sheet.append(['id', 'name'])
        sheet.append([1, 'A'])
        sheet.append([2, 'B'])
        sheet.append([3, 'C'])
        workbook.save(excel_path)

        with open(excel_path, 'rb') as f:
            files = {'file': ('bounds.xlsx', f.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}

        preflight = client.post(
            '/api/v1/datasource/preflight',
            files=files,
            data={'start_row': '0', 'start_col': '0', 'end_col': '1'},
        )
        assert preflight.status_code == 200
        preflight_id = preflight.json()['preflight_id']

        confirm_data = {
            'preflight_id': preflight_id,
            'name': 'Excel End Row',
            'sheet_name': 'Sheet1',
            'start_row': '0',
            'start_col': '0',
            'end_col': '1',
            'end_row': '1',
            'has_header': 'true',
        }
        confirm = client.post('/api/v1/datasource/confirm', data=confirm_data)

        assert confirm.status_code == 200
        config = confirm.json()['config']
        assert config['source']['end_row'] == 1

    def test_confirm_excel_cell_range_stores_bounds(self, client, temp_upload_dir: Path):
        """Test Excel confirm stores manual cell range selection."""
        excel_path = temp_upload_dir / 'cell-range.xlsx'
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            pytest.skip('Excel support not available')
        sheet.title = 'Sheet1'
        sheet.append(['id', 'name'])
        sheet.append([1, 'A'])
        sheet.append([2, 'B'])
        workbook.save(excel_path)

        with open(excel_path, 'rb') as f:
            files = {'file': ('cell-range.xlsx', f.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}

        preflight = client.post('/api/v1/datasource/preflight', files=files)
        assert preflight.status_code == 200
        preflight_id = preflight.json()['preflight_id']

        confirm_data = {
            'preflight_id': preflight_id,
            'name': 'Excel Range',
            'cell_range': 'Sheet1!A1:B3',
            'has_header': 'true',
        }
        confirm = client.post('/api/v1/datasource/confirm', data=confirm_data)

        assert confirm.status_code == 200
        config = confirm.json()['config']
        assert config['source']['cell_range'] == 'Sheet1!A1:B3'
        assert config['source']['sheet_name'] == 'Sheet1'
        assert config['source']['start_row'] == 0
        assert config['source']['start_col'] == 0
        assert config['source']['end_col'] == 1
        assert config['source']['end_row'] == 2

    @patch('modules.datasource.routes.run_db')
    def test_confirm_excel_preserves_validation_error(self, mock_run_db, client, temp_upload_dir: Path):
        excel_path = temp_upload_dir / 'confirm-invalid.xlsx'
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            pytest.skip('Excel support not available')
        sheet.title = 'Sheet1'
        sheet.append(['id', 'name'])
        sheet.append([1, 'A'])
        workbook.save(excel_path)

        with open(excel_path, 'rb') as f:
            files = {'file': ('confirm-invalid.xlsx', f.read(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}

        preflight = client.post('/api/v1/datasource/preflight', files=files)
        assert preflight.status_code == 200
        preflight_id = preflight.json()['preflight_id']

        mock_run_db.side_effect = DataSourceValidationError('Excel selection is invalid')
        confirm = client.post(
            '/api/v1/datasource/confirm',
            data={'preflight_id': preflight_id, 'name': 'Broken Excel', 'sheet_name': 'Sheet1', 'has_header': 'true'},
        )

        assert confirm.status_code == 400
        assert confirm.json()['detail'] == 'Excel selection is invalid'


class TestDataSourceSchema:
    """Test datasource schema extraction."""

    def test_schema_caching(self, client, test_db_session, sample_datasource: DataSource):
        """Test that schema is cached after first request."""
        # First request
        response1 = client.get(f'/api/v1/datasource/{sample_datasource.id}/schema')
        assert response1.status_code == 200

        # Check that schema_cache is populated
        test_db_session.refresh(sample_datasource)

        # Second request should use cache
        response2 = client.get(f'/api/v1/datasource/{sample_datasource.id}/schema')
        assert response2.status_code == 200

        # Responses should be identical
        assert response1.json() == response2.json()

    def test_schema_for_different_types(self, client, sample_csv_file: Path, sample_parquet_file: Path):
        """Test schema extraction for different file types."""
        # Upload CSV
        with open(sample_csv_file, 'rb') as f:
            files = {'file': ('test.csv', f.read(), 'text/csv')}
            data = {'name': 'CSV Schema Test'}
            response = client.post('/api/v1/datasource/upload', files=files, data=data)
            csv_id = response.json()['id']

        # Upload Parquet
        with open(sample_parquet_file, 'rb') as f:
            files = {'file': ('test.parquet', f.read(), 'application/octet-stream')}
            data = {'name': 'Parquet Schema Test'}
            response = client.post('/api/v1/datasource/upload', files=files, data=data)
            parquet_id = response.json()['id']

        # Get schemas
        csv_schema = client.get(f'/api/v1/datasource/{csv_id}/schema')
        parquet_schema = client.get(f'/api/v1/datasource/{parquet_id}/schema')

        assert csv_schema.status_code == 200
        assert parquet_schema.status_code == 200

        # Both should have columns
        assert 'columns' in csv_schema.json() or 'fields' in csv_schema.json()
        assert 'columns' in parquet_schema.json() or 'fields' in parquet_schema.json()


class TestDataSourceListing:
    """Test datasource listing functionality."""

    def test_list_empty_datasources(self, client):
        """Test listing when no datasources exist."""
        response = client.get('/api/v1/datasource')

        assert response.status_code == 200
        data = response.json()
        assert isinstance(data, list)

    def test_list_multiple_datasources(self, client, sample_datasources: list[DataSource]):
        """Test listing multiple datasources."""
        response = client.get('/api/v1/datasource')

        assert response.status_code == 200
        data = response.json()
        assert len(data) >= len(sample_datasources)

    def test_list_datasources_pagination(self, client, sample_datasources: list[DataSource]):
        """Test datasource listing with pagination."""
        # Test if pagination is supported
        response = client.get('/api/v1/datasource?limit=1')

        assert response.status_code == 200
        data = response.json()

        # Should return at most 1 item if pagination is supported
        if isinstance(data, list):
            assert len(data) <= 1 or len(data) == len(sample_datasources)

    def test_list_does_not_extract_or_write_schema_cache(self, client, test_db_session, sample_datasource: DataSource, monkeypatch):
        """Listing datasources stays read-only for schema cache."""
        calls = {'count': 0}

        def fail_extract(*args, **kwargs):
            calls['count'] += 1
            raise AssertionError('list_datasources must not call _extract_schema')

        monkeypatch.setattr('modules.datasource.service._extract_schema', fail_extract)

        response = client.get('/api/v1/datasource')
        assert response.status_code == 200
        assert calls['count'] == 0

        test_db_session.refresh(sample_datasource)
        assert sample_datasource.schema_cache is None


class TestComputeHistogram:
    """Test _compute_histogram helper function."""

    def test_empty_series(self):
        """Empty series returns empty list."""
        s = pl.Series('x', [], dtype=pl.Float64)
        result = _compute_histogram(s)
        assert result == []

    def test_single_value(self):
        """Single value collapses to one bin."""
        s = pl.Series('x', [5.0, 5.0, 5.0])
        result = _compute_histogram(s)
        assert len(result) == 1
        assert result[0]['start'] == 5.0
        assert result[0]['end'] == 5.0
        assert result[0]['count'] == 3

    def test_two_values(self):
        """Two distinct values are split across bins."""
        s = pl.Series('x', [0.0, 10.0])
        result = _compute_histogram(s, bins=2)
        assert len(result) == 2
        assert result[0]['start'] == 0.0
        assert result[1]['end'] == 10.0
        total = sum(b['count'] for b in result)  # type: ignore[call-overload]
        assert total == 2

    def test_default_20_bins(self):
        """Default call returns 20 bins."""
        s = pl.Series('x', list(range(100)))
        result = _compute_histogram(s)
        assert len(result) == 20

    def test_custom_bin_count(self):
        """Custom bins parameter is respected."""
        s = pl.Series('x', list(range(100)))
        result = _compute_histogram(s, bins=5)
        assert len(result) == 5

    def test_all_values_covered(self):
        """Sum of bin counts equals series length."""
        s = pl.Series('x', [1.0, 2.0, 3.0, 4.0, 5.0, 6.0, 7.0, 8.0, 9.0, 10.0])
        result = _compute_histogram(s, bins=5)
        total = sum(b['count'] for b in result)  # type: ignore[call-overload]
        assert total == 10

    def test_bins_are_contiguous(self):
        """Each bin starts where the previous one ended."""
        s = pl.Series('x', list(range(50)))
        result = _compute_histogram(s, bins=10)
        for i in range(1, len(result)):
            assert result[i]['start'] == result[i - 1]['end']

    def test_negative_values(self):
        """Handles negative ranges correctly."""
        s = pl.Series('x', [-10.0, -5.0, 0.0, 5.0, 10.0])
        result = _compute_histogram(s, bins=4)
        assert len(result) == 4
        assert result[0]['start'] == -10.0
        assert result[-1]['end'] == 10.0
        total = sum(b['count'] for b in result)  # type: ignore[call-overload]
        assert total == 5

    def test_integer_series(self):
        """Works with integer (non-float) series."""
        s = pl.Series('x', [1, 2, 3, 4, 5])
        result = _compute_histogram(s, bins=3)
        assert len(result) == 3
        total = sum(b['count'] for b in result)  # type: ignore[call-overload]
        assert total == 5

    def test_values_rounded(self):
        """Bin edges are rounded to 4 decimal places."""
        s = pl.Series('x', [0.0, 1.0 / 3.0, 2.0 / 3.0, 1.0])
        result = _compute_histogram(s, bins=3)
        for b in result:
            start_str = str(b['start'])
            if '.' in start_str:
                assert len(start_str.split('.')[1]) <= 4

    def test_single_element(self):
        """Single element series returns one bin."""
        s = pl.Series('x', [42.0])
        result = _compute_histogram(s)
        assert len(result) == 1
        assert result[0]['count'] == 1

    def test_with_nulls_filtered_out(self):
        """Histogram is computed on non-null values (caller should filter)."""
        s = pl.Series('x', [1.0, 2.0, 3.0])  # caller filters nulls before calling
        result = _compute_histogram(s, bins=3)
        total = sum(b['count'] for b in result)  # type: ignore[call-overload]
        assert total == 3


class TestDataSourceDeletion:
    """Test datasource deletion."""

    def test_delete_datasource_cascades(self, client, test_db_session, sample_analysis):
        """Test that deleting a datasource handles linked analyses."""
        datasource_id = sample_analysis.pipeline_definition['tabs'][0]['datasource']['id']

        # Delete the datasource
        response = client.delete(f'/api/v1/datasource/{datasource_id}')

        # Should succeed or return appropriate error
        assert response.status_code in [204, 400, 409]

    def test_delete_uploaded_datasource_removes_iceberg_and_upload(self, client):
        payload = b'id,value\n1,10\n2,20\n'
        files = {'file': ('to_delete.csv', payload, 'text/csv')}
        data = {'name': 'Delete Test'}

        create = client.post('/api/v1/datasource/upload', files=files, data=data)

        assert create.status_code == 200
        body = create.json()
        datasource_id = body['id']
        metadata_path = Path(body['config']['metadata_path'])
        source = body['config']['source']
        upload_path = Path(source['file_path'])

        assert metadata_path.exists()
        assert upload_path.exists()
        assert source['source_type'] == 'file'

        response = client.delete(f'/api/v1/datasource/{datasource_id}')

        assert response.status_code == 204
        assert not metadata_path.exists()
        assert not upload_path.exists()

        get_response = client.get(f'/api/v1/datasource/{datasource_id}')
        assert get_response.status_code == 404


class TestIsHidden:
    """Test is_hidden field on datasources."""

    def _insert_datasource(self, session, *, is_hidden: bool, csv_file: Path) -> DataSource:
        """Insert a datasource directly into the DB with the given visibility."""
        ds = DataSource(
            id=str(uuid.uuid4()),
            name='Hidden DS' if is_hidden else 'Visible DS',
            source_type='file',
            config={'file_path': str(csv_file), 'file_type': 'csv', 'options': {}},
            is_hidden=is_hidden,
            created_at=datetime.now(UTC),
        )
        session.add(ds)
        session.commit()
        session.refresh(ds)
        return ds

    def test_list_excludes_hidden_by_default(self, client, test_db_session, sample_csv_file: Path):
        """GET /api/v1/datasource omits hidden datasources by default."""
        visible = self._insert_datasource(test_db_session, is_hidden=False, csv_file=sample_csv_file)
        self._insert_datasource(test_db_session, is_hidden=True, csv_file=sample_csv_file)

        response = client.get('/api/v1/datasource')
        assert response.status_code == 200

        ids = [ds['id'] for ds in response.json()]
        assert visible.id in ids
        # Hidden datasource must not appear
        hidden_ids = [ds['id'] for ds in response.json() if ds.get('is_hidden')]
        assert hidden_ids == []

    def test_list_includes_hidden_when_requested(self, client, test_db_session, sample_csv_file: Path):
        """GET /api/v1/datasource?include_hidden=true returns both hidden and visible."""
        visible = self._insert_datasource(test_db_session, is_hidden=False, csv_file=sample_csv_file)
        hidden = self._insert_datasource(test_db_session, is_hidden=True, csv_file=sample_csv_file)

        response = client.get('/api/v1/datasource?include_hidden=true')
        assert response.status_code == 200

        ids = [ds['id'] for ds in response.json()]
        assert visible.id in ids
        assert hidden.id in ids

    def test_create_analysis_datasource_hidden(self, test_db_session, sample_analysis: Analysis):
        """create_analysis_datasource with is_hidden=True sets the flag."""
        result = create_analysis_datasource(
            test_db_session,
            name='Hidden Analysis DS',
            description=None,
            analysis_id=sample_analysis.id,
            is_hidden=True,
        )
        assert result.is_hidden is True

    def test_create_analysis_datasource_visible_by_default(self, test_db_session, sample_analysis: Analysis):
        """create_analysis_datasource without is_hidden defaults to False."""
        result = create_analysis_datasource(
            test_db_session,
            name='Visible Analysis DS',
            description=None,
            analysis_id=sample_analysis.id,
        )
        assert result.is_hidden is False

    def test_hidden_datasource_still_accessible_by_id(self, client, test_db_session, sample_csv_file: Path):
        """GET /api/v1/datasource/{id} returns a hidden datasource."""
        hidden = self._insert_datasource(test_db_session, is_hidden=True, csv_file=sample_csv_file)

        response = client.get(f'/api/v1/datasource/{hidden.id}')
        assert response.status_code == 200
        assert response.json()['id'] == hidden.id
        assert response.json()['is_hidden'] is True

    def test_auto_creation_on_analysis_update(self, client, test_db_session, sample_analysis: Analysis):
        """Updating an analysis keeps explicit datasource ids."""
        new_tab_id = str(uuid.uuid4())
        update_payload: dict[str, object] = {
            'tabs': [
                {
                    'id': new_tab_id,
                    'name': 'Auto Tab',
                    'parent_id': None,
                    'datasource': {
                        'id': sample_analysis.pipeline_definition['tabs'][0]['datasource']['id'],
                        'analysis_tab_id': None,
                        'config': {'branch': 'master'},
                    },
                    'output': {
                        'result_id': str(uuid.uuid4()),
                        'datasource_type': 'iceberg',
                        'format': 'parquet',
                        'filename': 'source_datasource',
                    },
                    'steps': [],
                },
            ],
        }

        response = client.put(f'/api/v1/analysis/{sample_analysis.id}', json=update_payload)
        assert response.status_code == 200


class TestDatasourceRefresh:
    @patch('datasource_service.load_datasource')
    @patch('datasource_service._write_iceberg_table')
    def test_refresh_external_datasource_updates_snapshot_fields(
        self,
        mock_write,
        mock_load,
        test_db_session,
        sample_csv_file: Path,
    ):
        from datasource_service import refresh_external_datasource

        class _Snap:
            snapshot_id = 222
            timestamp_ms = 654321

        class _Table:
            def current_snapshot(self):
                return _Snap()

        mock_load.return_value = pl.DataFrame({'x': [1]}).lazy()
        mock_write.return_value = _Table()

        ds = DataSource(
            id=str(uuid.uuid4()),
            name='Refreshable Raw',
            source_type='iceberg',
            config={
                'metadata_path': str(Path('data') / 'clean' / str(uuid.uuid4()) / 'master'),
                'branch': 'master',
                'source': {'source_type': 'file', 'file_path': str(sample_csv_file), 'file_type': 'csv', 'options': {}},
            },
            created_by='import',
            created_at=datetime.now(UTC),
        )
        test_db_session.add(ds)
        test_db_session.commit()

        out = refresh_external_datasource(test_db_session, ds.id)
        assert out.config['snapshot_id'] == '222'
        assert out.config['snapshot_timestamp_ms'] == 654321
        assert out.config['current_snapshot_id'] == '222'
        assert out.config['current_snapshot_timestamp_ms'] == 654321
        assert out.config['refresh'] is not None
        mock_write.assert_called_once_with(mock_load.return_value, Path(out.config['metadata_path']), build_mode='full')

        runs = (
            test_db_session.execute(select(EngineRun).where(EngineRun.datasource_id == ds.id))  # type: ignore[arg-type]
            .scalars()
            .all()
        )
        latest = sorted(runs, key=lambda row: row.created_at)[-1]
        assert latest.kind == 'datasource_update'
        assert latest.result_json is not None
        assert latest.result_json['snapshot_id'] == '222'


class TestDatasourceUpdateRunLogging:
    def test_update_raw_iceberg_logs_snapshot_id(self, client, test_db_session, sample_csv_file: Path):
        ds = DataSource(
            id=str(uuid.uuid4()),
            name='Raw Iceberg',
            source_type='iceberg',
            config={
                'metadata_path': str(Path('data') / 'clean' / str(uuid.uuid4()) / 'master'),
                'branch': 'master',
                'snapshot_id': '500',
                'source': {'source_type': 'file', 'file_path': str(sample_csv_file), 'file_type': 'csv', 'options': {}},
            },
            created_by='import',
            created_at=datetime.now(UTC),
        )
        test_db_session.add(ds)
        test_db_session.commit()

        response = client.put(f'/api/v1/datasource/{ds.id}', json={'name': 'Renamed Raw'})
        assert response.status_code == 200

        runs = (
            test_db_session.execute(select(EngineRun).where(EngineRun.datasource_id == ds.id))  # type: ignore[arg-type]
            .scalars()
            .all()
        )
        latest = sorted(runs, key=lambda row: row.created_at)[-1]
        assert latest.kind == 'datasource_update'
        assert latest.result_json is not None
        assert latest.result_json['snapshot_id'] == '500'

    def test_update_rejects_system_snapshot_fields(self, client, test_db_session, sample_csv_file: Path):
        ds = DataSource(
            id=str(uuid.uuid4()),
            name='Snapshot Locked',
            source_type='iceberg',
            config={
                'metadata_path': str(Path('data') / 'clean' / str(uuid.uuid4()) / 'master'),
                'branch': 'master',
                'snapshot_id': '111',
                'snapshot_timestamp_ms': 1000,
                'current_snapshot_id': '111',
                'current_snapshot_timestamp_ms': 1000,
                'source': {'source_type': 'file', 'file_path': str(sample_csv_file), 'file_type': 'csv', 'options': {}},
            },
            created_by='import',
            created_at=datetime.now(UTC),
        )
        test_db_session.add(ds)
        test_db_session.commit()

        response = client.put(
            f'/api/v1/datasource/{ds.id}',
            json={'config': {'current_snapshot_id': '999'}},
        )

        assert response.status_code == 400
        assert 'system-managed' in response.json()['detail']
