import asyncio
import contextvars
import hashlib
import json
import os
import sqlite3
from collections import deque
from collections.abc import Callable
from enum import StrEnum
from pathlib import Path
from threading import Lock
from urllib.parse import unquote, urlparse

import polars as pl
from iceberg_reader import scan_iceberg_snapshot
from openpyxl import load_workbook
from pydantic import ConfigDict
from step_converter import convert_step_format

from contracts.compute.base import OperationHandler, OperationParams
from core.namespace import namespace_paths


class DatasourceSourceType(StrEnum):
    FILE = 'file'
    DATABASE = 'database'
    DUCKDB = 'duckdb'
    ICEBERG = 'iceberg'
    ANALYSIS = 'analysis'


class IcebergReader(StrEnum):
    NATIVE = 'native'
    PYICEBERG = 'pyiceberg'


class IcebergMetadataPathNotFoundError(ValueError):
    def __init__(self, metadata_path: str):
        self.metadata_path = metadata_path
        super().__init__(f'Iceberg metadata_path not found: {metadata_path}')


class DatasourceParams(OperationParams):
    model_config = ConfigDict(extra='allow')

    source_type: DatasourceSourceType = DatasourceSourceType.FILE
    analysis_tab_id: str | None = None
    analysis_pipeline: dict | None = None
    file_path: str | None = None
    file_type: str | None = None
    options: dict | None = None
    csv_options: dict | None = None
    sheet_name: str | None = None
    start_row: int | None = None
    start_col: int | None = None
    end_col: int | None = None
    end_row: int | None = None
    has_header: bool | None = None
    table_name: str | None = None
    named_range: str | None = None
    cell_range: str | None = None
    connection_string: str | None = None
    query: str | None = None
    db_path: str | None = None
    read_only: bool = True
    metadata_path: str | None = None
    branch: str | None = None
    namespace_name: str | None = None
    snapshot_id: str | None = None
    snapshot_timestamp_ms: int | None = None
    storage_options: dict | None = None
    reader: IcebergReader | None = None


class DatasourceHandler(OperationHandler):
    FILE_LOADERS: dict[str, Callable[[str, dict], pl.LazyFrame]] = {}
    SOURCE_LOADERS: dict[str, Callable[['DatasourceHandler', DatasourceParams], pl.LazyFrame]] = {}

    def __call__(
        self,
        lf: pl.LazyFrame,
        params: dict,
        **_,
    ) -> pl.LazyFrame:
        validated = DatasourceParams.model_validate(params)
        loader = self.SOURCE_LOADERS.get(validated.source_type.value)
        if not loader:
            allowed = ', '.join(sorted(self.SOURCE_LOADERS))
            raise ValueError(f'Unsupported source type: {validated.source_type}. Allowed: {allowed}')
        return loader(self, validated)

    async def call_async(
        self,
        lf: pl.LazyFrame,
        params: dict,
        **_,
    ) -> pl.LazyFrame:
        return await asyncio.to_thread(self.__call__, lf, params)

    @staticmethod
    def _csv_opts(opts: dict | None) -> dict:
        if not opts:
            return {}
        return {
            'separator': opts.get('delimiter', ','),
            'quote_char': opts.get('quote_char', '"'),
            'has_header': opts.get('has_header', True),
            'skip_rows': opts.get('skip_rows', 0),
            'encoding': opts.get('encoding', 'utf8'),
            'try_parse_dates': True,
        }

    def _load_file(self, config: DatasourceParams) -> pl.LazyFrame:
        if not config.file_path or not config.file_type:
            raise ValueError('Datasource file loading requires file_path and file_type')
        if config.file_type == 'excel' and _has_bounds(config):
            return _read_excel_bounds(config)
        opts = config.csv_options or config.options or {}
        opts = _merge_excel_opts(config, opts)
        loader = self.FILE_LOADERS.get(config.file_type)
        if not loader:
            raise ValueError(f'Unsupported file type: {config.file_type}')
        return loader(config.file_path, opts)

    def _load_database(self, config: DatasourceParams) -> pl.LazyFrame:
        if not config.connection_string or not config.query:
            raise ValueError('Datasource database loading requires connection_string and query')
        _assert_select_only(config.query)
        if config.connection_string.startswith('sqlite:'):
            parsed = urlparse(config.connection_string)
            if not parsed.path:
                raise ValueError('SQLite connection string must include a database path')
            connection = sqlite3.connect(parsed.path)
            try:
                frame = pl.read_database(config.query, connection)
            except Exception as exc:
                raise ValueError('Failed to query database datasource') from exc
            finally:
                connection.close()
            return frame.lazy()
        return pl.read_database_uri(config.query, config.connection_string).lazy()

    def _load_duckdb(self, config: DatasourceParams) -> pl.LazyFrame:
        import duckdb

        if not config.query:
            raise ValueError('Datasource DuckDB loading requires query')
        _assert_select_only(config.query)
        conn = (
            duckdb.connect(database=config.db_path, read_only=config.read_only) if config.db_path else duckdb.connect(database=':memory:')
        )
        try:
            return conn.execute(config.query).fetch_df().lazy()
        finally:
            conn.close()

    def _load_iceberg(self, config: DatasourceParams) -> pl.LazyFrame:
        if not config.metadata_path:
            raise ValueError('Datasource Iceberg loading requires metadata_path')
        metadata_path = resolve_iceberg_branch_metadata_path(
            config.metadata_path,
            config.branch,
            namespace_name=config.namespace_name,
        )
        snapshot_id = config.snapshot_id
        snapshot_value: int | None = None
        if snapshot_id is not None:
            try:
                snapshot_value = int(snapshot_id)
            except (TypeError, ValueError) as exc:
                raise ValueError(f'Iceberg snapshot ID must be an integer: {snapshot_id}') from exc
        if snapshot_id is None and config.snapshot_timestamp_ms is not None:
            from pyiceberg.table import StaticTable

            table = StaticTable.from_metadata(metadata_path)
            snapshot = table.snapshot_as_of_timestamp(config.snapshot_timestamp_ms)
            if snapshot is None:
                raise ValueError('Iceberg snapshot not found for the selected timestamp')
            snapshot_value = snapshot.snapshot_id
        reader_override = IcebergReader.NATIVE if config.reader == IcebergReader.NATIVE else IcebergReader.PYICEBERG
        if snapshot_value is not None:
            return scan_iceberg_snapshot(metadata_path, snapshot_value, config.storage_options)
        return pl.scan_iceberg(
            metadata_path,
            snapshot_id=snapshot_value,
            storage_options=config.storage_options,
            reader_override=reader_override.value,
        )

    def _load_analysis(self, config: DatasourceParams) -> pl.LazyFrame:
        pipeline = config.analysis_pipeline
        if isinstance(pipeline, dict):
            pipeline_id = pipeline.get('analysis_id')
            if pipeline_id:
                return _load_analysis_pipeline(pipeline, str(pipeline_id), config.analysis_tab_id)
        raise ValueError('analysis_pipeline is required for analysis datasource loading')


_analysis_stack_var: contextvars.ContextVar[tuple[tuple[str, str | None], ...]] = contextvars.ContextVar(
    '_analysis_stack',
    default=(),
)
_ANALYSIS_CACHE: dict[str, pl.LazyFrame] = {}
_ANALYSIS_CACHE_KEYS: deque[str] = deque()
_ANALYSIS_CACHE_LOCK = Lock()
_ANALYSIS_CACHE_MAX = 20


def _get_analysis_stack() -> tuple[tuple[str, str | None], ...]:
    return _analysis_stack_var.get()


def _load_analysis_pipeline(pipeline: dict, analysis_id: str, tab_id: str | None) -> pl.LazyFrame:
    stack = _get_analysis_stack()
    stack_key = (analysis_id, tab_id)
    if stack_key in stack:
        raise ValueError('Analysis pipeline contains a circular reference')
    token = _analysis_stack_var.set((*stack, stack_key))
    try:
        cache_key = _analysis_cache_key(pipeline, tab_id)
        cached = _get_analysis_cache(cache_key)
        if cached is not None:
            return cached
        frame = _build_analysis_from_pipeline(pipeline, tab_id)
        _store_analysis_cache(cache_key, frame)
        return frame
    finally:
        _analysis_stack_var.reset(token)


def _analysis_cache_key(pipeline: dict, tab_id: str | None) -> str:
    payload = {
        'analysis_id': pipeline.get('analysis_id'),
        'tab_id': tab_id,
        'tabs': pipeline.get('tabs', []),
    }
    data = json.dumps(payload, sort_keys=True, default=str)
    return hashlib.sha256(data.encode('utf-8')).hexdigest()


def _resolve_pipeline_datasource(pipeline: dict, datasource: dict) -> dict:
    datasource_id = datasource.get('id')
    if not isinstance(datasource_id, str) or not datasource_id:
        raise ValueError('Analysis tab missing datasource.id')
    config = datasource.get('config')
    if not isinstance(config, dict):
        raise ValueError('Analysis tab datasource.config must be a dict')
    branch = config.get('branch')
    if not isinstance(branch, str) or not branch.strip():
        raise ValueError('Analysis tab datasource.config.branch is required')
    analysis_tab_id = datasource.get('analysis_tab_id')
    if isinstance(analysis_tab_id, str) and analysis_tab_id:
        return {
            'source_type': 'analysis',
            'analysis_id': str(pipeline.get('analysis_id') or ''),
            'analysis_tab_id': analysis_tab_id,
            **config,
        }
    source_type = datasource.get('source_type')
    if not isinstance(source_type, str) or not source_type.strip():
        raise ValueError(f'Analysis pipeline missing datasource metadata for {datasource_id}')
    return {'source_type': source_type, **config}


def _step_source_payload(step_config: dict, source_id: str) -> dict | None:
    direct = step_config.get('right_source_datasource')
    if isinstance(direct, dict) and direct.get('id') == source_id:
        return direct
    many = step_config.get('source_datasources')
    if not isinstance(many, list):
        return None
    return next(
        (item for item in many if isinstance(item, dict) and isinstance(item.get('id'), str) and item.get('id') == source_id),
        None,
    )


def _get_analysis_cache(key: str) -> pl.LazyFrame | None:
    with _ANALYSIS_CACHE_LOCK:
        return _ANALYSIS_CACHE.get(key)


def _store_analysis_cache(key: str, frame: pl.LazyFrame) -> None:
    with _ANALYSIS_CACHE_LOCK:
        if key in _ANALYSIS_CACHE:
            return
        _ANALYSIS_CACHE[key] = frame
        _ANALYSIS_CACHE_KEYS.append(key)
        if len(_ANALYSIS_CACHE_KEYS) <= _ANALYSIS_CACHE_MAX:
            return
        oldest = _ANALYSIS_CACHE_KEYS.popleft()
        _ANALYSIS_CACHE.pop(oldest, None)


def _resolve_analysis_tab(tabs: list[dict], analysis_tab_id: str | None) -> dict:
    selected = None
    if analysis_tab_id:
        selected = next((tab for tab in tabs if tab.get('id') == analysis_tab_id), None)
    if not selected:
        selected = next((tab for tab in tabs if tab.get('steps')), None)
    if not selected:
        raise ValueError('Analysis pipeline missing tab steps')
    return selected


def _resolve_tab_chain(tabs: list[dict], target_tab_id: str) -> list[dict]:
    output_to_tab: dict[str, dict] = {}
    tab_input: dict[str, str] = {}
    output_map: dict[str, str] = {}
    for tab in tabs:
        tab_id = tab.get('id')
        output = tab.get('output') if isinstance(tab, dict) else None
        output_id = output.get('result_id') if isinstance(output, dict) else None
        if not tab_id or not output_id:
            continue
        output_map[str(tab_id)] = str(output_id)
    for tab in tabs:
        tid = tab.get('id')
        output = tab.get('output') if isinstance(tab, dict) else None
        output_id = output.get('result_id') if isinstance(output, dict) else None
        datasource = tab.get('datasource') if isinstance(tab, dict) else None
        input_id = datasource.get('id') if isinstance(datasource, dict) else None
        if tid and output_id:
            output_to_tab[str(output_id)] = tab
        if tid and input_id:
            tab_input[str(tid)] = str(input_id)

    ordered: list[dict] = []
    seen: set[str] = set()
    current_id = target_tab_id
    while current_id:
        if current_id in seen:
            break
        seen.add(current_id)
        current_tab = next((tab for tab in tabs if tab.get('id') == current_id), None)
        if not current_tab:
            break
        ordered.append(current_tab)
        input_ds = tab_input.get(current_id)
        if input_ds and input_ds in output_to_tab:
            upstream_tab = output_to_tab[input_ds]
            current_id = str(upstream_tab.get('id')) if upstream_tab.get('id') else ''
            continue
        break
    ordered.reverse()
    return ordered


def _build_tab_pipeline(
    tab: dict,
    pipeline: dict,
    cache: dict[str, pl.LazyFrame],
) -> pl.LazyFrame:
    datasource = tab.get('datasource') if isinstance(tab, dict) else None
    if not isinstance(datasource, dict):
        raise ValueError('Analysis tab datasource must be a dict')
    datasource_id = datasource.get('id')
    if not datasource_id:
        raise ValueError('Analysis tab missing datasource.id')

    if datasource_id in cache:
        base_frame = cache[datasource_id]
    else:
        merged = _resolve_pipeline_datasource(pipeline, datasource)
        analysis_id = pipeline.get('analysis_id')
        analysis_id = str(analysis_id) if analysis_id is not None else None
        if analysis_id and merged.get('source_type') == 'analysis' and str(merged.get('analysis_id')) == analysis_id:
            merged = {**merged, 'analysis_pipeline': pipeline}

        base_frame = load_datasource(merged)

    steps = tab.get('steps', [])
    if not isinstance(steps, list):
        raise ValueError('Analysis tab steps must be a list')

    from compute_utils import apply_steps

    steps = apply_steps(steps)
    additional = _collect_analysis_sources(steps, pipeline, cache)

    from compute_engine import PolarsComputeEngine

    for step in steps:
        step_id = step.get('id') or 'step'
        backend_step = convert_step_format(step)
        right_source_raw = backend_step.params.get('right_source')
        right_source_id = right_source_raw if isinstance(right_source_raw, str) else None
        right_lf = cache.get(right_source_id) if right_source_id is not None else None
        base_frame = PolarsComputeEngine._apply_step(
            base_frame,
            backend_step,
            right_sources=additional,
            right_lf=right_lf,
        )
        cache[step_id] = base_frame

    output = tab.get('output') if isinstance(tab, dict) else None
    output_id = output.get('result_id') if isinstance(output, dict) else None
    if output_id:
        cache[str(output_id)] = base_frame

    return base_frame


def _build_analysis_from_pipeline(pipeline: dict, analysis_tab_id: str | None) -> pl.LazyFrame:
    tabs = pipeline.get('tabs', [])
    if not isinstance(tabs, list) or not tabs:
        raise ValueError('Analysis pipeline missing tabs')

    selected = _resolve_analysis_tab(tabs, analysis_tab_id)
    target_id = selected.get('id') if selected else None
    if not target_id:
        raise ValueError('Analysis pipeline missing tab id')

    chain = _resolve_tab_chain(tabs, str(target_id))
    cache: dict[str, pl.LazyFrame] = {}
    last_frame: pl.LazyFrame | None = None
    for tab in chain:
        last_frame = _build_tab_pipeline(tab, pipeline, cache)

    if last_frame is None:
        raise ValueError('Analysis pipeline did not produce a LazyFrame')

    return last_frame


def _collect_analysis_sources(
    steps: list[dict],
    pipeline: dict,
    cache: dict[str, pl.LazyFrame],
) -> dict[str, pl.LazyFrame]:
    additional: dict[str, pl.LazyFrame] = {}
    output_to_tab = {}
    tabs = pipeline.get('tabs', [])
    if isinstance(tabs, list):
        for tab in tabs:
            if not isinstance(tab, dict):
                continue
            tab_id = tab.get('id')
            output = tab.get('output')
            output_id = output.get('result_id') if isinstance(output, dict) else None
            if isinstance(tab_id, str) and isinstance(output_id, str) and output_id:
                output_to_tab[output_id] = tab_id
    for step in steps:
        config = step.get('config', {})
        if not isinstance(config, dict):
            continue
        right_source_id = config.get('right_source') or config.get('rightDataSource')

        union_sources = config.get('sources', [])
        if isinstance(union_sources, str):
            union_sources = [union_sources]

        source_ids = ([right_source_id] if right_source_id else []) + union_sources

        for source_id in source_ids:
            if not source_id:
                continue
            if str(source_id) in cache:
                additional[str(source_id)] = cache[str(source_id)]
                continue
            if str(source_id) in output_to_tab:
                next_config = {
                    'source_type': 'analysis',
                    'analysis_id': str(pipeline.get('analysis_id') or ''),
                    'analysis_tab_id': output_to_tab[str(source_id)],
                    'analysis_pipeline': pipeline,
                }
            else:
                raw = _step_source_payload(config, str(source_id))
                if not isinstance(raw, dict):
                    raise ValueError(f'Analysis pipeline missing datasource config for {source_id}')
                next_config = _resolve_pipeline_datasource(pipeline, raw)
                if next_config.get('source_type') == 'analysis':
                    next_config = {**next_config, 'analysis_pipeline': pipeline}
            additional[str(source_id)] = load_datasource(next_config)

    return additional


def resolve_iceberg_metadata_path(
    metadata_path: str,
    *,
    namespace_name: str | None = None,
    data_root: str | Path | None = None,
) -> str:
    normalized = _strip_file_scheme(metadata_path)
    path = Path(normalized)
    resolved = path.resolve()
    root = _resolve_iceberg_data_root(namespace_name=namespace_name, data_root=data_root)
    if root not in resolved.parents and root != resolved:
        raise ValueError('Iceberg metadata_path must be inside data directory')
    if path.suffix == '.db':
        raise ValueError('Iceberg metadata_path must point to metadata.json, not catalog.db')
    if path.is_file():
        if path.name.endswith('.metadata.json'):
            return str(path)
        raise ValueError('Iceberg metadata_path must be a table directory or metadata.json')
    if not path.exists():
        raise IcebergMetadataPathNotFoundError(metadata_path)
    if not path.is_dir():
        raise ValueError(f'Iceberg metadata_path must be a file or directory: {metadata_path}')
    if path.name == 'metadata':
        return _latest_metadata_file(path)
    metadata_dir = path / 'metadata'
    if metadata_dir.is_dir():
        return _latest_metadata_file(metadata_dir)
    raise ValueError('Iceberg metadata_path must be a table directory containing metadata/')


def resolve_iceberg_branch_metadata_path(
    metadata_path: str,
    branch: str | None,
    *,
    namespace_name: str | None = None,
    data_root: str | Path | None = None,
) -> str:
    normalized = _strip_file_scheme(metadata_path)
    path = Path(normalized)
    if path.suffix == '.metadata.json' or path.name == 'metadata' or path.is_file():
        return resolve_iceberg_metadata_path(metadata_path, namespace_name=namespace_name, data_root=data_root)
    if branch:
        branch_path = path / branch
        if branch_path.exists():
            return resolve_iceberg_metadata_path(str(branch_path), namespace_name=namespace_name, data_root=data_root)
    metadata_dir = path / 'metadata'
    if metadata_dir.is_dir():
        return resolve_iceberg_metadata_path(str(metadata_dir), namespace_name=namespace_name, data_root=data_root)
    if path.is_dir():
        children = [entry for entry in path.iterdir() if entry.is_dir()]
        if len(children) == 1:
            return resolve_iceberg_metadata_path(str(children[0]), namespace_name=namespace_name, data_root=data_root)
    return resolve_iceberg_metadata_path(metadata_path, namespace_name=namespace_name, data_root=data_root)


def _resolve_iceberg_data_root(*, namespace_name: str | None = None, data_root: str | Path | None = None) -> Path:
    if data_root is not None:
        return Path(os.path.realpath(Path(data_root).resolve()))
    return Path(os.path.realpath(namespace_paths(namespace_name).base_dir.resolve()))


def _read_excel(path: str, opts: dict) -> pl.LazyFrame:
    sheet_name = opts.get('sheet_name')
    table_name = opts.get('table_name')
    named_range = opts.get('named_range')
    next_opts: dict = {}
    if sheet_name is not None:
        next_opts['sheet_name'] = sheet_name
    if table_name is not None:
        next_opts['table_name'] = table_name
    if named_range is not None:
        next_opts['named_range'] = named_range
    return pl.read_excel(path, **next_opts).lazy()


def _merge_excel_opts(config: DatasourceParams, opts: dict) -> dict:
    next_opts = opts
    if config.sheet_name:
        next_opts = {**next_opts, 'sheet_name': config.sheet_name}
    if config.table_name:
        next_opts = {**next_opts, 'table_name': config.table_name}
    if config.named_range:
        next_opts = {**next_opts, 'named_range': config.named_range}
    if config.cell_range:
        next_opts = {**next_opts, 'cell_range': config.cell_range}
    if config.has_header is not None:
        next_opts = {**next_opts, 'has_header': config.has_header}
    return next_opts


def _read_excel_bounds(config: DatasourceParams) -> pl.LazyFrame:
    file_path = config.file_path
    sheet_name = config.sheet_name
    start_row = config.start_row
    start_col = config.start_col
    end_row = config.end_row
    end_col = config.end_col
    has_header = config.has_header if config.has_header is not None else True
    if not file_path or start_row is None or start_col is None or end_row is None or end_col is None:
        raise ValueError('Excel bounds require file_path, start_row, start_col, end_row, end_col')

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    target_sheet = sheet_name or workbook.sheetnames[0]
    sheet = workbook[target_sheet]
    rows = list(
        sheet.iter_rows(
            min_row=start_row + 1,
            max_row=end_row + 1,
            min_col=start_col + 1,
            max_col=end_col + 1,
            values_only=True,
        ),
    )

    if not rows:
        return pl.DataFrame().lazy()

    if has_header:
        header = rows[0]
        columns = _normalize_headers(header)
        data_rows = rows[1:]
    else:
        columns = [f'column_{index + 1}' for index in range(len(rows[0]))]
        data_rows = rows
    frame = pl.DataFrame(data_rows, schema=columns, orient='row')
    return frame.lazy()


DatasourceHandler.FILE_LOADERS = {
    'csv': lambda path, opts: pl.scan_csv(path, **DatasourceHandler._csv_opts(opts)),
    'parquet': lambda path, _: pl.scan_parquet(path),
    'json': lambda path, _: pl.read_json(path).lazy(),
    'ndjson': lambda path, _: pl.scan_ndjson(path),
    'excel': _read_excel,
}

DatasourceHandler.SOURCE_LOADERS = {
    'file': DatasourceHandler._load_file,
    'database': DatasourceHandler._load_database,
    'duckdb': DatasourceHandler._load_duckdb,
    'iceberg': DatasourceHandler._load_iceberg,
    'analysis': DatasourceHandler._load_analysis,
}


def _latest_metadata_file(metadata_dir: Path) -> str:
    candidates = list(metadata_dir.glob('*.metadata.json'))
    if not candidates:
        raise ValueError(f'No metadata.json files found in {metadata_dir}')
    return str(max(candidates, key=_metadata_sort_key))


def _strip_file_scheme(metadata_path: str) -> str:
    if not metadata_path.startswith('file:'):
        return metadata_path
    parsed = urlparse(metadata_path)
    if parsed.scheme != 'file':
        return metadata_path
    if parsed.netloc and parsed.netloc != 'localhost':
        if parsed.path:
            return unquote(f'/{parsed.netloc}{parsed.path}')
        return unquote(f'/{parsed.netloc}')
    if parsed.path:
        return unquote(parsed.path)
    return metadata_path


def _metadata_sort_key(path: Path) -> tuple[int, int, float]:
    name = path.name
    prefix = name.split('-', maxsplit=1)[0]
    if prefix.isdigit():
        return (1, int(prefix), path.stat().st_mtime)
    return (0, 0, path.stat().st_mtime)


def _normalize_headers(values: tuple[object | None, ...]) -> list[str]:
    names: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values):
        base = str(value).strip() if value is not None else f'column_{index + 1}'
        if base not in seen:
            seen[base] = 0
            names.append(base)
            continue
        seen[base] += 1
        names.append(f'{base}_{seen[base]}')
    return names


def _has_bounds(config: DatasourceParams) -> bool:
    return config.start_row is not None and config.start_col is not None and config.end_col is not None and config.end_row is not None


def _assert_select_only(query: str) -> None:
    """Reject queries that are not read-only SELECT statements."""
    first_token = query.strip().split()[0].upper() if query.strip() else ''
    if first_token not in ('SELECT', 'WITH'):
        raise ValueError('Only SELECT queries (including CTEs starting with WITH) are permitted for database datasources')


def load_datasource(config: dict) -> pl.LazyFrame:
    return DatasourceHandler()(pl.LazyFrame(), config)


async def load_datasource_async(config: dict) -> pl.LazyFrame:
    return await asyncio.to_thread(load_datasource, config)
