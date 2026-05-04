from __future__ import annotations

from enum import StrEnum
from typing import Any

import polars as pl
import psycopg
from openpyxl import load_workbook

from core.iceberg_metadata import resolve_iceberg_branch_metadata_path
from core.iceberg_snapshot_reader import scan_iceberg_snapshot


class DatasourceSourceType(StrEnum):
    FILE = 'file'
    DATABASE = 'database'
    DUCKDB = 'duckdb'
    ICEBERG = 'iceberg'


class IcebergReader(StrEnum):
    NATIVE = 'native'
    PYICEBERG = 'pyiceberg'


def _csv_opts(opts: dict[str, Any] | None) -> dict[str, Any]:
    if not opts:
        return {}
    return {
        'separator': opts.get('delimiter', ','),
        'quote_char': opts.get('quote_char', '"'),
        'has_header': opts.get('has_header', True),
        'skip_rows': opts.get('skip_rows', 0),
        'encoding': opts.get('encoding', 'utf8'),
        'try_parse_dates': True,
    }


def _read_excel(path: str, opts: dict[str, Any]) -> pl.LazyFrame:
    sheet_name = opts.get('sheet_name')
    table_name = opts.get('table_name')
    named_range = opts.get('named_range')
    next_opts: dict[str, Any] = {}
    if sheet_name is not None:
        next_opts['sheet_name'] = sheet_name
    if table_name is not None:
        next_opts['table_name'] = table_name
    if named_range is not None:
        next_opts['named_range'] = named_range
    return pl.read_excel(path, **next_opts).lazy()


def _merge_excel_opts(config: dict[str, Any], opts: dict[str, Any]) -> dict[str, Any]:
    next_opts = dict(opts)
    if config.get('sheet_name'):
        next_opts['sheet_name'] = config['sheet_name']
    if config.get('table_name'):
        next_opts['table_name'] = config['table_name']
    if config.get('named_range'):
        next_opts['named_range'] = config['named_range']
    if config.get('cell_range'):
        next_opts['cell_range'] = config['cell_range']
    if config.get('has_header') is not None:
        next_opts['has_header'] = config['has_header']
    return next_opts


def _normalize_headers(values: tuple[object | None, ...]) -> list[str]:
    names: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values):
        base = str(value).strip() if value is not None else f'column_{index + 1}'
        if base not in seen:
            seen[base] = 0
            names.append(base)
            continue
        seen[base] += 1
        names.append(f'{base}_{seen[base]}')
    return names


def _has_bounds(config: dict[str, Any]) -> bool:
    return (
        config.get('start_row') is not None
        and config.get('start_col') is not None
        and config.get('end_col') is not None
        and config.get('end_row') is not None
    )


def _read_excel_bounds(config: dict[str, Any]) -> pl.LazyFrame:
    file_path = config.get('file_path')
    sheet_name = config.get('sheet_name')
    start_row = config.get('start_row')
    start_col = config.get('start_col')
    end_row = config.get('end_row')
    end_col = config.get('end_col')
    has_header = config.get('has_header') if config.get('has_header') is not None else True
    if not file_path or start_row is None or start_col is None or end_row is None or end_col is None:
        raise ValueError('Excel bounds require file_path, start_row, start_col, end_row, end_col')

    workbook = load_workbook(file_path, read_only=True, data_only=True)
    target_sheet = sheet_name or workbook.sheetnames[0]
    sheet = workbook[target_sheet]
    rows = list(
        sheet.iter_rows(
            min_row=int(start_row) + 1,
            max_row=int(end_row) + 1,
            min_col=int(start_col) + 1,
            max_col=int(end_col) + 1,
            values_only=True,
        ),
    )

    if not rows:
        return pl.DataFrame().lazy()

    if has_header:
        header = rows[0]
        columns = _normalize_headers(header)
        data_rows = rows[1:]
    else:
        columns = [f'column_{index + 1}' for index in range(len(rows[0]))]
        data_rows = rows
    frame = pl.DataFrame(data_rows, schema=columns, orient='row')
    return frame.lazy()


def _assert_select_only(query: str) -> None:
    first_token = query.strip().split()[0].upper() if query.strip() else ''
    if first_token not in ('SELECT', 'WITH'):
        raise ValueError('Only SELECT queries (including CTEs starting with WITH) are permitted for database datasources')


def load_datasource_frame(config: dict[str, Any]) -> pl.LazyFrame:
    source_type = str(config.get('source_type') or '')
    if source_type == DatasourceSourceType.FILE.value:
        file_path = config.get('file_path')
        file_type = config.get('file_type')
        if not file_path or not file_type:
            raise ValueError('Datasource file loading requires file_path and file_type')
        if file_type == 'excel' and _has_bounds(config):
            return _read_excel_bounds(config)
        opts = config.get('csv_options') or config.get('options') or {}
        if not isinstance(opts, dict):
            opts = {}
        opts = _merge_excel_opts(config, opts)
        if file_type == 'csv':
            return pl.scan_csv(file_path, **_csv_opts(opts))
        if file_type == 'parquet':
            return pl.scan_parquet(file_path)
        if file_type == 'json':
            return pl.read_json(file_path).lazy()
        if file_type == 'ndjson':
            return pl.scan_ndjson(file_path)
        if file_type == 'excel':
            return _read_excel(file_path, opts)
        raise ValueError(f'Unsupported file type: {file_type}')

    if source_type == DatasourceSourceType.DATABASE.value:
        connection_string = config.get('connection_string')
        query = config.get('query')
        if not isinstance(connection_string, str) or not isinstance(query, str):
            raise ValueError('Datasource database loading requires connection_string and query')
        _assert_select_only(query)
        if not connection_string.lower().startswith('postgresql://'):
            raise ValueError('Database datasource connection string must be PostgreSQL')
        with psycopg.connect(connection_string, autocommit=True) as connection:
            return pl.read_database(query, connection).lazy()

    if source_type == DatasourceSourceType.DUCKDB.value:
        import duckdb

        query = config.get('query')
        if not isinstance(query, str):
            raise ValueError('Datasource DuckDB loading requires query')
        _assert_select_only(query)
        db_path = config.get('db_path')
        read_only = bool(config.get('read_only', True))
        conn = duckdb.connect(database=db_path, read_only=read_only) if db_path else duckdb.connect(database=':memory:')
        try:
            return conn.execute(query).fetch_df().lazy()
        finally:
            conn.close()

    if source_type == DatasourceSourceType.ICEBERG.value:
        metadata_path = config.get('metadata_path')
        if not isinstance(metadata_path, str):
            raise ValueError('Datasource Iceberg loading requires metadata_path')
        resolved_metadata_path = resolve_iceberg_branch_metadata_path(
            metadata_path,
            config.get('branch') if isinstance(config.get('branch'), str) else None,
            namespace_name=config.get('namespace_name') if isinstance(config.get('namespace_name'), str) else None,
        )
        snapshot_id = config.get('snapshot_id')
        snapshot_value: int | None = None
        if snapshot_id is not None:
            try:
                snapshot_value = int(snapshot_id)
            except (TypeError, ValueError) as exc:
                raise ValueError(f'Iceberg snapshot ID must be an integer: {snapshot_id}') from exc
        if snapshot_value is not None:
            storage_options = config.get('storage_options')
            return scan_iceberg_snapshot(
                resolved_metadata_path,
                snapshot_value,
                storage_options if isinstance(storage_options, dict) else None,
            )
        reader = config.get('reader')
        reader_override = IcebergReader.NATIVE if reader == IcebergReader.NATIVE.value else IcebergReader.PYICEBERG
        storage_options = config.get('storage_options')
        return pl.scan_iceberg(
            resolved_metadata_path,
            snapshot_id=snapshot_value,
            storage_options=storage_options if isinstance(storage_options, dict) else None,
            reader_override=reader_override.value,
        )

    raise ValueError(f'Unsupported source type: {source_type}')
